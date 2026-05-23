"""
Étape 16 — Pondération du capital par score de régularité par symbole

Idée : au lieu d'allouer 5% fixe à chaque symbole, allouer proportionnellement
à la régularité historique de chaque symbole individuel.

Étapes :
  1. Simuler chaque symbole ISOLÉ sur chaque année 2018-2024 → score de régularité
  2. Convertir les scores en poids de capital (avec floor 2%)
  3. Simuler sur 4 ans avec ces poids via sim_multi_weighted()
  4. Comparer : 5% fixe vs pondéré avec floor vs pondéré sans floor

Config : Top 20 / 12h / épurée / multi / sansSL (meilleure config validée)
"""

import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

import config
import data_cache
import fear_greed
import multi_sim as ms
from multi_sim import sim_multi_weighted

init(autoreset=True)

SYMBOLS   = config.SYMBOLS
TIMEFRAME = "12h"
SCORE_YEARS = list(range(2018, 2025))   # 2018-2024 pour calculer la régularité
MIN_YEARS   = 4                          # minimum d'années pour avoir un score fiable
FLOOR_PCT   = 0.02                       # chaque symbole reçoit au moins 2%

# ---------------------------------------------------------------------------
# Étape 1 — Régularité par symbole
# ---------------------------------------------------------------------------
def compute_symbol_regularity(symbols: list[str], tf: str, years: list[int],
                               fg: dict) -> dict:
    """
    Pour chaque symbole, simule en isolation sur chaque année calendaire
    et calcule : moy_%, pct_positif, score_reg = moy × (pct_pos/100)²
    """
    print(f"\n{Fore.CYAN}Calcul régularité par symbole ({len(symbols)} symboles × {len(years)} années)...{Style.RESET_ALL}")
    regularite = {}
    total = len(symbols) * len(years)
    done  = 0

    for sym in symbols:
        returns = []
        for year in years:
            df_year = ms.get_df_for_year(sym, tf, year)
            if df_year is None or len(df_year) < 20:
                done += 1
                continue
            try:
                r = ms.sim_multi_on_dfs(
                    {sym: df_year},
                    use_sl=False,
                    fg=fg,
                    use_triple_st=False,
                    use_sma_macd=False,
                    tf=tf,
                )
                ret = r.get("return_%")
                if ret is not None:
                    returns.append((year, ret))
            except Exception:
                pass
            done += 1
            print(f"\r  {done}/{total}", end="", flush=True)

        if len(returns) >= MIN_YEARS:
            vals    = [r for _, r in returns]
            moy     = sum(vals) / len(vals)
            pct_pos = sum(1 for v in vals if v > 0) / len(vals)
            score   = moy * (pct_pos ** 2)
            regularite[sym] = {
                "score":    round(score, 2),
                "moy_%":   round(moy, 1),
                "pct_pos": round(pct_pos * 100, 0),
                "n_années": len(returns),
                "détail":   returns,
            }
        else:
            # Pas assez d'historique → score neutre (floor sera appliqué)
            regularite[sym] = {
                "score":    0.0,
                "moy_%":   0.0,
                "pct_pos": 0.0,
                "n_années": len(returns),
                "détail":   returns,
            }

    print()
    return regularite


# ---------------------------------------------------------------------------
# Étape 2 — Conversion en poids de capital
# ---------------------------------------------------------------------------
def scores_to_weights(regularite: dict, symbols: list[str],
                      floor_pct: float = 0.0) -> dict[str, float]:
    """
    Transforme les scores de régularité en poids de capital par symbole.
    floor_pct : allocation minimum garantie (ex. 0.02 = 2%)
    La somme des poids est normalisée pour que le total déployable max = 100%.
    """
    # Plancher à 0 pour les symboles négatifs
    scores_pos = {s: max(regularite[s]["score"], 0.0) for s in symbols}
    total_score = sum(scores_pos.values())

    if total_score == 0:
        # Tous négatifs → poids égaux
        return {s: 1.0 / len(symbols) for s in symbols}

    # Poids proportionnels aux scores
    poids_raw = {s: scores_pos[s] / total_score for s in symbols}

    if floor_pct == 0:
        return poids_raw

    # Application du floor : chaque symbole reçoit au moins floor_pct
    # On redistribue le surplus proportionnellement
    n = len(symbols)
    total_floor = floor_pct * n

    if total_floor >= 1.0:
        # Floor trop élevé → égalité forcée
        return {s: 1.0 / n for s in symbols}

    remaining = 1.0 - total_floor
    # Poids de base = floor + part proportionnelle du surplus
    poids = {s: floor_pct + poids_raw[s] * remaining for s in symbols}
    return poids


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_period_dfs(symbols, tf, years_back):
    n = ms.CANDLES_PER_YEAR[tf] * years_back
    dfs = {}
    for sym in symbols:
        df = ms.get_df(sym, tf)
        if df is not None and len(df) >= 10:
            dfs[sym] = df.tail(n).reset_index(drop=True)
    return dfs


def fmt(v):
    if v is None:
        return "—"
    color = Fore.GREEN if v > 0 else (Fore.RED if v < 0 else "")
    return f"{color}{v:+.1f}%{Style.RESET_ALL}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{Fore.CYAN}{'='*80}")
    print("  ÉTAPE 16 — Pondération du capital par régularité symbole")
    print(f"  Config : Top 20 / {TIMEFRAME} / épurée / multi / sansSL")
    print(f"{'='*80}{Style.RESET_ALL}")

    # ---- Chargement données ------------------------------------------------
    print(f"\n{Fore.CYAN}Chargement données...{Style.RESET_ALL}")
    data_cache.prefetch_all(SYMBOLS, [TIMEFRAME], verbose=False)
    data_cache.prefetch_all_8y(SYMBOLS, [TIMEFRAME], verbose=False)
    data_cache.prefetch_all_10y(SYMBOLS, [TIMEFRAME], verbose=False)
    for sym in SYMBOLS:
        ms.get_df(sym, TIMEFRAME)
        ms.get_df_8y(sym, TIMEFRAME)
        ms.get_df_10y(sym, TIMEFRAME)
    print("  OK")

    fg_data = fear_greed.load(verbose=False)
    print(f"  Fear & Greed : {len(fg_data)} jours")

    # ---- Étape 1 : régularité par symbole ----------------------------------
    regularite = compute_symbol_regularity(SYMBOLS, TIMEFRAME, SCORE_YEARS, fg_data)

    # Affichage du tableau de régularité
    print(f"\n{Fore.YELLOW}{'='*80}")
    print("  RÉGULARITÉ PAR SYMBOLE (2018-2024, config épurée / 12h / sansSL)")
    print(f"{'='*80}{Style.RESET_ALL}")

    reg_headers = ["Symbole", "Score", "Moy%/an", "Pct+", "N années"] + [str(y) for y in SCORE_YEARS]
    reg_rows = []
    for sym in sorted(regularite, key=lambda s: regularite[s]["score"], reverse=True):
        r = regularite[sym]
        detail_map = dict(r["détail"])
        row = [
            sym.replace("/USDT", ""),
            f"{Fore.CYAN}{r['score']:+.1f}{Style.RESET_ALL}",
            fmt(r["moy_%"]),
            f"{r['pct_pos']:.0f}%",
            r["n_années"],
        ]
        for y in SCORE_YEARS:
            v = detail_map.get(y)
            row.append(fmt(v) if v is not None else "—")
        reg_rows.append(row)
    print(tabulate(reg_rows, headers=reg_headers, tablefmt="rounded_outline"))

    # ---- Étape 2 : conversion en poids -------------------------------------
    poids_avec_floor  = scores_to_weights(regularite, SYMBOLS, floor_pct=FLOOR_PCT)
    poids_sans_floor  = scores_to_weights(regularite, SYMBOLS, floor_pct=0.0)

    # Affichage des poids
    print(f"\n{Fore.YELLOW}{'='*80}")
    print(f"  POIDS DE CAPITAL (floor={FLOOR_PCT*100:.0f}% vs sans floor vs 5% fixe)")
    print(f"{'='*80}{Style.RESET_ALL}")

    poids_headers = ["Symbole", "Score", "5% fixe", "Pondéré (floor 2%)", "Pondéré (sans floor)"]
    poids_rows = []
    for sym in sorted(regularite, key=lambda s: poids_avec_floor[s], reverse=True):
        pf  = poids_avec_floor[sym]
        psf = poids_sans_floor[sym]
        arrow_f  = "▲" if pf  > 0.05 else ("▼" if pf  < 0.04 else "≈")
        arrow_sf = "▲" if psf > 0.05 else ("▼" if psf < 0.04 else "≈")
        col_f  = Fore.GREEN if pf  > 0.05 else (Fore.RED if pf  < 0.04 else "")
        col_sf = Fore.GREEN if psf > 0.05 else (Fore.RED if psf < 0.04 else "")
        poids_rows.append([
            sym.replace("/USDT", ""),
            f"{regularite[sym]['score']:+.1f}",
            "5.0%",
            f"{col_f}{pf*100:.1f}% {arrow_f}{Style.RESET_ALL}",
            f"{col_sf}{psf*100:.1f}% {arrow_sf}{Style.RESET_ALL}",
        ])
    print(tabulate(poids_rows, headers=poids_headers, tablefmt="rounded_outline"))

    # ---- Étape 3 : simulations comparatives --------------------------------
    print(f"\n{Fore.CYAN}Simulations comparatives (4 périodes × 3 configs)...{Style.RESET_ALL}")

    configs_sim = [
        ("5% fixe (réf)",         None,              None),
        ("Pondéré + floor 2%",    poids_avec_floor,  True),
        ("Pondéré sans floor",    poids_sans_floor,  False),
    ]

    period_headers = ["Config", "1 an", "2 ans", "3 ans", "4 ans",
                      "Trades(4a)", "Win%(4a)", "DD%(4a)"]
    period_rows = []

    for label, poids, has_floor in configs_sim:
        row = [label]
        r4  = {}
        for years in [1, 2, 3, 4]:
            dfs = load_period_dfs(SYMBOLS, TIMEFRAME, years)
            if poids is None:
                r = ms.sim_multi_on_dfs(dfs, use_sl=False, fg=fg_data,
                                        use_triple_st=False, use_sma_macd=False,
                                        tf=TIMEFRAME)
            else:
                r = sim_multi_weighted(dfs, pos_pct_per_symbol=poids,
                                       fg=fg_data,
                                       use_triple_st=False, use_sma_macd=False,
                                       use_sl=False, tf=TIMEFRAME)
            row.append(fmt(r.get("return_%")))
            if years == 4:
                r4 = r
        row.extend([
            r4.get("trades", "—"),
            f"{r4.get('win_%', 0):.1f}%",
            f"{r4.get('drawdown_%', 0):.1f}%",
        ])
        period_rows.append(row)

    print(f"\n{Fore.YELLOW}{'='*80}")
    print("  COMPARAISON PAR PÉRIODE")
    print(f"{'='*80}{Style.RESET_ALL}")
    print(tabulate(period_rows, headers=period_headers, tablefmt="rounded_outline"))

    # ---- Étape 4 : par année calendaire ------------------------------------
    import datetime
    current_year = datetime.date.today().year
    cal_years = list(range(2018, current_year + 1))

    print(f"\n{Fore.YELLOW}{'='*80}")
    print("  COMPARAISON PAR ANNÉE CALENDAIRE")
    print(f"{'='*80}{Style.RESET_ALL}")

    year_headers = ["Config"] + [str(y) for y in cal_years] + ["Pct+", "Moy/an"]
    year_rows = []

    for label, poids, _ in configs_sim:
        row     = [label]
        returns = []
        for year in cal_years:
            dfs = {}
            for sym in SYMBOLS:
                df_year = ms.get_df_for_year(sym, TIMEFRAME, year)
                if df_year is not None:
                    dfs[sym] = df_year
            if not dfs:
                row.append("—")
                continue
            if poids is None:
                r = ms.sim_multi_on_dfs(dfs, use_sl=False, fg=fg_data,
                                        use_triple_st=False, use_sma_macd=False,
                                        tf=TIMEFRAME)
            else:
                r = sim_multi_weighted(dfs, pos_pct_per_symbol=poids,
                                       fg=fg_data,
                                       use_triple_st=False, use_sma_macd=False,
                                       use_sl=False, tf=TIMEFRAME)
            ret = r.get("return_%")
            if ret is not None:
                returns.append(ret)
                row.append(fmt(ret))
            else:
                row.append("—")

        pct_pos = f"{sum(1 for v in returns if v > 0) / len(returns) * 100:.0f}%" if returns else "—"
        moy     = f"{sum(returns)/len(returns):+.1f}%" if returns else "—"
        row.extend([pct_pos, moy])
        year_rows.append(row)

    print(tabulate(year_rows, headers=year_headers, tablefmt="rounded_outline"))

    # ---- Résumé final -------------------------------------------------------
    print(f"\n{Fore.YELLOW}{'='*80}")
    print("  RÉSUMÉ — Impact de la pondération")
    print(f"{'='*80}{Style.RESET_ALL}")

    sum_headers = ["Config", "4 ans", "Trades", "Win%", "DD%", "Verdict"]
    sum_rows    = []
    ref_ret     = None

    for label, poids, _ in configs_sim:
        dfs = load_period_dfs(SYMBOLS, TIMEFRAME, 4)
        if poids is None:
            r = ms.sim_multi_on_dfs(dfs, use_sl=False, fg=fg_data,
                                    use_triple_st=False, use_sma_macd=False,
                                    tf=TIMEFRAME)
        else:
            r = sim_multi_weighted(dfs, pos_pct_per_symbol=poids,
                                   fg=fg_data,
                                   use_triple_st=False, use_sma_macd=False,
                                   use_sl=False, tf=TIMEFRAME)
        ret = r.get("return_%", 0)
        if ref_ret is None:
            ref_ret = ret
        delta = ret - ref_ret
        if poids is None:
            verdict = "référence"
        elif delta > 2:
            verdict = f"{Fore.GREEN}+{delta:+.1f}% vs réf{Style.RESET_ALL}"
        elif delta < -2:
            verdict = f"{Fore.RED}{delta:+.1f}% vs réf{Style.RESET_ALL}"
        else:
            verdict = f"≈ réf ({delta:+.1f}%)"

        color = Fore.GREEN if ret > 0 else Fore.RED
        sum_rows.append([
            label,
            f"{color}{ret:+.2f}%{Style.RESET_ALL}",
            r.get("trades", "—"),
            f"{r.get('win_%', 0):.1f}%",
            f"{r.get('drawdown_%', 0):.1f}%",
            verdict,
        ])

    print(tabulate(sum_rows, headers=sum_headers, tablefmt="rounded_outline"))
    print()

    # ---- Export Excel -------------------------------------------------------
    xlsx_path = "etape16_ponderation_regularite.xlsx"
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GREEN_H = PatternFill("solid", fgColor="C6EFCE")
        GREY_H  = PatternFill("solid", fgColor="D9D9D9")
        BOLD    = Font(bold=True)
        CENTER  = Alignment(horizontal="center")
        thin    = Side(style="thin", color="CCCCCC")
        BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_sheet(ws, n_rows, n_cols, green_cols=None):
            green_cols = green_cols or set()
            for col_idx in range(1, n_cols + 1):
                h = ws.cell(row=1, column=col_idx)
                h.fill = GREEN_H if col_idx in green_cols else GREY_H
                h.font = BOLD
                h.alignment = CENTER
                h.border = BORDER
                for row_idx in range(2, n_rows + 2):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.alignment = CENTER
                    c.border = BORDER
            ws.freeze_panes = "A2"

        # Feuille 1 — Régularité par symbole
        reg_export = []
        for sym in sorted(regularite, key=lambda s: regularite[s]["score"], reverse=True):
            r = regularite[sym]
            detail_map = dict(r["détail"])
            row_d = {
                "symbole":  sym,
                "score":    r["score"],
                "moy_%":   r["moy_%"],
                "pct_pos": r["pct_pos"],
                "n_années": r["n_années"],
            }
            for y in SCORE_YEARS:
                row_d[str(y)] = detail_map.get(y)
            row_d["poids_fixe_%"]        = 5.0
            row_d["poids_floor2_%"]      = round(poids_avec_floor[sym] * 100, 2)
            row_d["poids_sans_floor_%"]  = round(poids_sans_floor[sym] * 100, 2)
            reg_export.append(row_d)

        df_reg = pd.DataFrame(reg_export)

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_reg.to_excel(writer, index=False, sheet_name="Régularité par symbole")
            ws = writer.sheets["Régularité par symbole"]
            for col_idx, col in enumerate(df_reg.columns, start=1):
                max_len = max(len(str(col)), df_reg[col].astype(str).str.len().max())
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 18)
            style_sheet(ws, len(df_reg), len(df_reg.columns), green_cols={2, 3, 4})

        print(f"{Fore.GREEN}Export → {xlsx_path}{Style.RESET_ALL}")
    except ImportError:
        print(f"{Fore.YELLOW}openpyxl non installé — Excel ignoré{Style.RESET_ALL}")
