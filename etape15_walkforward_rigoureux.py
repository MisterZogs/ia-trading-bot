"""
Étape 15 — Walk-Forward rigoureux : split fixe train / test

  Train : 2017-2023 (7 ans d'optimisation)
  Test  : 2024-2025 (hors-sample complet)

Différence avec etape8 (8 fenêtres glissantes) :
  - Un seul split clair, non touché pendant le développement de la stratégie
  - 2024-2025 = données jamais utilisées pour aucun réglage → vrai verdict OOS

Configs testées :
  Portfolios  : top20, top5
  TF          : 12h
  Stratégies  : épurée, baseline, +TripleST, épurée+ST
  SL          : avecSL, sansSL
  Capital     : 5%/illim, 10%/max10
"""

import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

import config
import data_cache
import fear_greed
import multi_sim as ms
from multi_sim import sim_concentration

init(autoreset=True)

# ---------------------------------------------------------------------------
# Paramètres
# ---------------------------------------------------------------------------
TIMEFRAME   = "12h"
TRAIN_YEARS = list(range(2017, 2024))   # 2017-2023
TEST_YEARS  = [2024, 2025]

PORTFOLIOS = {
    "Top 20": config.SYMBOLS,
    "Top 5":  ["BTC/USDT", "ETH/USDT", "BNB/USDT", "SOL/USDT", "XRP/USDT"],
}

STRATEGIES = [
    ("épurée",    False, False),   # use_triple_st, use_sma_macd
    ("baseline",  False, True),
    ("+TripleST", True,  True),
    ("épurée+ST", True,  False),
]

CAPITAL_SCHEMES = [
    ("5%/illim",  0.05, 9999),
    ("10%/max10", 0.10,   10),
]

SL_OPTIONS = [
    ("sansSL", False),
    ("avecSL", True),
]

INITIAL_CAPITAL = ms.INITIAL_CAPITAL

# ---------------------------------------------------------------------------
# Construction des DFs train / test
# ---------------------------------------------------------------------------
def build_period_dfs(symbols: list[str], tf: str, years: list[int]) -> dict:
    """
    Concatène les DFs année par année pour former la période complète.
    Chaque année est chargée via get_df_for_year() qui choisit le bon tier
    de données (4y / 8y / 10y) automatiquement.
    """
    result = {}
    for sym in symbols:
        parts = []
        for year in years:
            df_year = ms.get_df_for_year(sym, tf, year)
            if df_year is not None and len(df_year) > 0:
                parts.append(df_year)
        if parts:
            result[sym] = pd.concat(parts, ignore_index=True)
    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def fmt(v, fmt_str="+.1f"):
    if v is None:
        return "—"
    color = Fore.GREEN if v > 0 else (Fore.RED if v < 0 else "")
    return f"{color}{v:{fmt_str}}%{Style.RESET_ALL}"


def verdict(r_train, r_test):
    if r_train is None or r_test is None:
        return "—"
    ratio = r_test / r_train if r_train != 0 else 0
    if r_test > 0 and ratio >= 0.5:
        return f"{Fore.GREEN}Robuste{Style.RESET_ALL}"
    elif r_test > 0 and ratio >= 0.2:
        return f"{Fore.YELLOW}Acceptable{Style.RESET_ALL}"
    elif r_test <= 0 and r_train > 10:
        return f"{Fore.RED}Overfitting{Style.RESET_ALL}"
    elif r_test <= 0:
        return f"{Fore.RED}Négatif OOS{Style.RESET_ALL}"
    else:
        return f"{Fore.YELLOW}Mitigé{Style.RESET_ALL}"


def run_sim(dfs, use_sl, use_triple_st, use_sma_macd, pos_pct, max_trades, fg, tf):
    if pos_pct != 0.05 or max_trades != 9999:
        return sim_concentration(dfs, pos_pct=pos_pct, max_trades=max_trades,
                                 fg=fg, use_triple_st=use_triple_st,
                                 use_sma_macd=use_sma_macd, use_sl=use_sl, tf=tf)
    return ms.sim_multi_on_dfs(dfs, use_sl=use_sl, fg=fg,
                                use_triple_st=use_triple_st, use_sma_macd=use_sma_macd,
                                tf=tf)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{Fore.CYAN}{'='*90}")
    print("  ÉTAPE 15 — Walk-Forward rigoureux")
    print(f"  Train : {TRAIN_YEARS[0]}–{TRAIN_YEARS[-1]}  |  Test OOS : {TEST_YEARS[0]}–{TEST_YEARS[-1]}")
    print(f"{'='*90}{Style.RESET_ALL}")

    # ---- Chargement données ------------------------------------------------
    all_symbols = list(dict.fromkeys(s for syms in PORTFOLIOS.values() for s in syms))
    print(f"\n{Fore.CYAN}Chargement des données historiques (8y + 10y)...{Style.RESET_ALL}")
    data_cache.prefetch_all_8y(all_symbols, [TIMEFRAME], verbose=False)
    data_cache.prefetch_all_10y(all_symbols, [TIMEFRAME], verbose=False)

    print(f"{Fore.CYAN}Calcul des indicateurs...{Style.RESET_ALL}")
    for sym in all_symbols:
        ms.get_df_8y(sym, TIMEFRAME)
        ms.get_df_10y(sym, TIMEFRAME)
    print("  OK")

    fg_data = fear_greed.load(verbose=False)
    print(f"  Fear & Greed : {len(fg_data)} jours")

    # ---- Construction des périodes -----------------------------------------
    print(f"\n{Fore.CYAN}Construction des DFs train ({TRAIN_YEARS[0]}-{TRAIN_YEARS[-1]}) et test ({TEST_YEARS})...{Style.RESET_ALL}")
    train_dfs_cache = {}
    test_dfs_cache  = {}
    for port_label, symbols in PORTFOLIOS.items():
        train_dfs_cache[port_label] = build_period_dfs(symbols, TIMEFRAME, TRAIN_YEARS)
        test_dfs_cache[port_label]  = build_period_dfs(symbols, TIMEFRAME, TEST_YEARS)
        n_train = len(train_dfs_cache[port_label])
        n_test  = len(test_dfs_cache[port_label])
        print(f"  {port_label} : {n_train} symboles train / {n_test} symboles test")

    # ---- Simulations --------------------------------------------------------
    rows = []
    configs = [
        (port, strat_name, use_triple_st, use_sma_macd, sl_label, use_sl, cap_label, pos_pct, max_t)
        for port in PORTFOLIOS
        for strat_name, use_triple_st, use_sma_macd in STRATEGIES
        for sl_label, use_sl in SL_OPTIONS
        for cap_label, pos_pct, max_t in CAPITAL_SCHEMES
    ]
    total = len(configs)
    print(f"\n{Fore.CYAN}Simulations ({total} configs × 2 périodes)...{Style.RESET_ALL}")

    for i, (port, strat, use_ts, use_sm, sl_lbl, use_sl, cap_lbl, pos_pct, max_t) in enumerate(configs):
        train_dfs = train_dfs_cache[port]
        test_dfs  = test_dfs_cache[port]

        try:
            r_train = run_sim(train_dfs, use_sl, use_ts, use_sm, pos_pct, max_t, fg_data, TIMEFRAME)
        except Exception:
            r_train = {}
        try:
            r_test = run_sim(test_dfs, use_sl, use_ts, use_sm, pos_pct, max_t, fg_data, TIMEFRAME)
        except Exception:
            r_test = {}

        ret_tr = r_train.get("return_%")
        ret_te = r_test.get("return_%")
        ratio  = round(ret_te / ret_tr, 2) if ret_tr and ret_tr != 0 and ret_te is not None else None

        rows.append({
            "portfolio": port,
            "stratégie": strat,
            "sl":        sl_lbl,
            "capital":   cap_lbl,
            "ret_train_%": ret_tr,
            "dd_train_%":  r_train.get("drawdown_%"),
            "trades_train": r_train.get("trades"),
            "ret_test_%":  ret_te,
            "dd_test_%":   r_test.get("drawdown_%"),
            "trades_test": r_test.get("trades"),
            "ratio":       ratio,
            "verdict":     (ret_tr, ret_te),
        })
        print(f"\r  {i+1}/{total}", end="", flush=True)
    print()

    # ---- Affichage principal ------------------------------------------------
    print(f"\n{Fore.YELLOW}{'='*120}")
    print("  RÉSULTATS — Train (2017-2023) vs Test OOS (2024-2025)")
    print(f"{'='*120}{Style.RESET_ALL}")

    headers = ["Portfolio", "Stratégie", "SL", "Capital",
               "Train%", "DD train%", "Trades train",
               "Test OOS%", "DD test%", "Trades test",
               "Ratio T/Tr", "Verdict"]
    table_rows = []
    for r in sorted(rows, key=lambda x: x["ret_test_%"] or -999, reverse=True):
        vrd = verdict(r["verdict"][0], r["verdict"][1])
        ratio_str = f"{r['ratio']:.2f}" if r["ratio"] is not None else "—"
        table_rows.append([
            r["portfolio"], r["stratégie"], r["sl"], r["capital"],
            fmt(r["ret_train_%"]),
            fmt(r["dd_train_%"]) if r["dd_train_%"] is not None else "—",
            r["trades_train"] or "—",
            fmt(r["ret_test_%"]),
            fmt(r["dd_test_%"]) if r["dd_test_%"] is not None else "—",
            r["trades_test"] or "—",
            ratio_str,
            vrd,
        ])
    print(tabulate(table_rows, headers=headers, tablefmt="rounded_outline"))

    # ---- Résumé par stratégie -----------------------------------------------
    print(f"\n{Fore.CYAN}{'='*70}")
    print("  RÉSUMÉ PAR STRATÉGIE (médiane test OOS, toutes configs)")
    print(f"{'='*70}{Style.RESET_ALL}")

    df_rows = pd.DataFrame(rows)
    if not df_rows.empty and df_rows["ret_test_%"].notna().any():
        for strat in df_rows["stratégie"].unique():
            sub = df_rows[df_rows["stratégie"] == strat]["ret_test_%"].dropna()
            if not sub.empty:
                med = sub.median()
                color = Fore.GREEN if med > 0 else Fore.RED
                bar = "█" * max(0, int(abs(med) / 3))
                print(f"  {strat:<15} {color}{med:+6.1f}%{Style.RESET_ALL}  {bar}")

    # ---- Verdict global -----------------------------------------------
    print(f"\n{Fore.CYAN}{'='*70}")
    print("  VERDICT GLOBAL")
    print(f"{'='*70}{Style.RESET_ALL}")

    n_robuste  = sum(1 for r in rows if r["ret_test_%"] is not None and r["ret_test_%"] > 0
                     and r["ratio"] is not None and r["ratio"] >= 0.5)
    n_positif  = sum(1 for r in rows if r["ret_test_%"] is not None and r["ret_test_%"] > 0)
    n_total    = sum(1 for r in rows if r["ret_test_%"] is not None)
    best = max(rows, key=lambda x: x["ret_test_%"] or -999)

    print(f"  Configs positives OOS  : {Fore.GREEN}{n_positif}/{n_total}{Style.RESET_ALL}")
    print(f"  Configs robustes OOS   : {Fore.GREEN}{n_robuste}/{n_total}{Style.RESET_ALL} (ratio ≥ 0.5)")
    if best["ret_test_%"] is not None:
        print(f"  Meilleure config OOS   : {best['portfolio']} / {best['stratégie']} / {best['sl']} / {best['capital']}")
        ratio_disp = f"{best['ratio']:.2f}" if best["ratio"] is not None else "—"
        print(f"    → Train: {fmt(best['ret_train_%'])}  |  Test OOS: {fmt(best['ret_test_%'])}  |  Ratio: {ratio_disp}")

    # ---- Export Excel -------------------------------------------------------
    xlsx_path = "etape15_walkforward_rigoureux.xlsx"
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GREEN_H = PatternFill("solid", fgColor="C6EFCE")
        BLUE_H  = PatternFill("solid", fgColor="BDD7EE")
        GREY_H  = PatternFill("solid", fgColor="D9D9D9")
        BOLD    = Font(bold=True)
        CENTER  = Alignment(horizontal="center")
        thin    = Side(style="thin", color="CCCCCC")
        BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

        export_rows = []
        for r in sorted(rows, key=lambda x: x["ret_test_%"] or -999, reverse=True):
            v0, v1 = r["verdict"]
            rat = r["ratio"]
            export_rows.append({
                "portfolio":    r["portfolio"],
                "stratégie":    r["stratégie"],
                "sl":           r["sl"],
                "capital":      r["capital"],
                "ret_train_%":  r["ret_train_%"],
                "dd_train_%":   r["dd_train_%"],
                "trades_train": r["trades_train"],
                "ret_test_%":   r["ret_test_%"],
                "dd_test_%":    r["dd_test_%"],
                "trades_test":  r["trades_test"],
                "ratio":        rat,
                "verdict":      "Robuste" if v0 and v1 and v1 > 0 and rat and rat >= 0.5
                                else "Positif OOS" if v1 and v1 > 0
                                else "Overfitting" if v0 and v0 > 10 and v1 is not None and v1 <= 0
                                else "Négatif OOS",
            })

        df_export = pd.DataFrame(export_rows)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Résultats")
            ws = writer.sheets["Résultats"]
            for col_idx, col_name in enumerate(df_export.columns, start=1):
                max_len = max(len(str(col_name)),
                              df_export[col_name].astype(str).str.len().max())
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 20)
                h = ws.cell(row=1, column=col_idx)
                h.fill = GREEN_H if "test" in col_name else (BLUE_H if "train" in col_name else GREY_H)
                h.font = BOLD
                h.alignment = CENTER
                h.border = BORDER
                for row_idx in range(2, len(df_export) + 2):
                    ws.cell(row=row_idx, column=col_idx).alignment = CENTER
                    ws.cell(row=row_idx, column=col_idx).border = BORDER
            ws.freeze_panes = "A2"
        print(f"\n{Fore.GREEN}Export → {xlsx_path}{Style.RESET_ALL}")
    except ImportError:
        print(f"{Fore.YELLOW}openpyxl non installé — Excel ignoré{Style.RESET_ALL}")
    print()
