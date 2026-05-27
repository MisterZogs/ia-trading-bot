"""
TP Ranking Full — toutes les dimensions de full_ranking.py + dimension TP.
Utilise multiprocessing (fork) pour paralléliser les simulations.

Dimensions :
  TP         : 2%, 3%, 5%, 10%, 15%, 20%, ATR dynamique, sans TP
  Portfolio  : top20, top10, top5, btceth, btc, eth
  Mode       : single (1pos/paire), multi (N pos/paire)
  SL         : avecSL, sansSL
  TF         : 30m, 1h, 2h, 4h, 6h, 12h, 1d
  Période    : 1an, 2ans, 3ans, 4ans + années calendaires 2018–2025
  Stratégie  : baseline, +TripleST, épurée, épurée+ST
  Capital    : 5%/illim, 10%/max10, 20%/max5, 50%/max2, 100%/max1

Usage :
    python3 tp_ranking_full.py
"""

import math
import os
import multiprocessing
import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

import config
import data_cache
import fear_greed
import indicators
import multi_sim as ms
from multi_sim import sim_concentration

init(autoreset=True)

# ---------------------------------------------------------------------------
# Dimensions
# ---------------------------------------------------------------------------
ALL_TF  = ["30m", "1h", "2h", "4h", "6h", "12h", "1d"]
OLD_TF  = data_cache.TF_WITH_8Y
PERIODS = {1: "1an", 2: "2ans", 3: "3ans", 4: "4ans"}

PORTFOLIOS = {
    "top20":  config.SYMBOLS,
    "top10":  ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT",
               "ADA/USDT","AVAX/USDT","DOT/USDT","LINK/USDT","MATIC/USDT"],
    "top5":   ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT"],
    "btceth": ["BTC/USDT","ETH/USDT"],
    "btc":    ["BTC/USDT"],
    "eth":    ["ETH/USDT"],
}
PORT_LABELS = {
    "top20":"Top 20","top10":"Top 10","top5":"Top 5",
    "btceth":"BTC+ETH","btc":"BTC","eth":"ETH",
}

CAPITAL_SCHEMES = [
    ("5%/illim",  0.05, 9999),
    ("10%/max10", 0.10,   10),
    ("20%/max5",  0.20,    5),
    ("50%/max2",  0.50,    2),
    ("100%/max1", 1.00,    1),
]

STRATEGIES = [
    ("baseline",  False, True,  3),
    ("+TripleST", True,  True,  3),
    ("épurée",    False, False, 3),
    ("épurée+ST", True,  False, 3),
]

TP_MODES = [
    ("2%",    True,  0.02,  False),
    ("3%",    True,  0.03,  False),
    ("5%",    True,  0.05,  False),
    ("10%",   True,  0.10,  False),
    ("15%",   True,  0.15,  False),
    ("20%",   True,  0.20,  False),
    ("ATR",   True,  None,  True),
    ("sansTP",False, None,  False),
]

CALENDAR_YEARS = list(range(2018, 2026))

# ---------------------------------------------------------------------------
# État global partagé via fork (peuplé dans le processus principal)
# ---------------------------------------------------------------------------
_SIG_CACHE: dict = {}   # key → (dfs, sigs)
_FG_DATA:   dict = {}
_BAH_CACHE: dict = {}

# ---------------------------------------------------------------------------
# Fonction worker — tourne dans les processus enfants (fork)
# ---------------------------------------------------------------------------
def _worker(task):
    (sig_key, port_label, mode_label, sl_label, tf, period_label,
     strat_name, min_score, cap_label, pos_pct, max_trades,
     tp_label, use_tp, tp_pct, atr_tp,
     bah, is_old, is_year) = task

    dfs, sigs = _SIG_CACHE.get(sig_key, ({}, {}))
    if not dfs:
        return None

    original = config.MIN_SCORE_TO_TRADE
    config.MIN_SCORE_TO_TRADE = min_score
    try:
        r = sim_concentration(
            dfs,
            pos_pct=pos_pct, max_trades=max_trades,
            fg=_FG_DATA,
            use_sl=(sl_label == "avecSL"),
            tf=tf,
            single=(mode_label == "1pos"),
            use_tp=use_tp, tp_pct=tp_pct, atr_tp=atr_tp,
            precomputed_sigs=sigs,
        )
    except Exception:
        r = {}
    finally:
        config.MIN_SCORE_TO_TRADE = original

    ret = r.get("return_%")
    if ret is None and is_year:
        return None

    nb = r.get("trades")
    alpha = round(ret - bah, 1) if ret is not None and bah is not None else None

    years = 1 if is_year else {"1an": 1, "2ans": 2, "3ans": 3, "4ans": 4}.get(period_label, 1)
    vol = round(nb * pos_pct / years, 1) if nb else None

    return {
        "tp":               tp_label,
        "portfolio":        port_label,
        "mode":             mode_label,
        "sl":               sl_label,
        "timeframe":        tf,
        "période":          period_label,
        "stratégie":        strat_name,
        "capital":          cap_label,
        "return_%":         ret,
        "drawdown_%":       r.get("drawdown_%"),
        "win_%":            r.get("win_%"),
        "trades":           nb,
        "pf":               r.get("profit_factor"),
        "dur.moy_j":        r.get("avg_duration_j"),
        "non_ferm.":        r.get("unclosed"),
        "bah_%":            bah,
        "alpha_%":          alpha,
        "volume_annuel_%":  vol,
        "is_old":           is_old,
        "sig_key":          sig_key,
    }

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _slice_recent(syms, tf, years):
    n = ms.CANDLES_PER_YEAR[tf] * years
    dfs = {}
    for sym in syms:
        df = ms.get_df(sym, tf)
        if df is not None and len(df) >= 10:
            dfs[sym] = df.tail(n).reset_index(drop=True)
    return dfs

def _slice_old(syms, tf, years):
    n   = ms.CANDLES_PER_YEAR[tf] * years
    n4y = ms.CANDLES_PER_YEAR[tf] * 4
    dfs = {}
    for sym in syms:
        df = ms.get_df_8y(sym, tf)
        if df is not None and len(df) > n4y:
            sliced = df.iloc[:-n4y].tail(n).reset_index(drop=True)
            if len(sliced) >= int(n * 0.9):
                dfs[sym] = sliced
    return dfs

def _slice_year(syms, tf, year):
    dfs = {}
    for sym in syms:
        df = ms.get_df_for_year(sym, tf, year)
        if df is not None and len(df) >= 10:
            dfs[sym] = df
    return dfs

def _compute_sigs(dfs, use_triple_st, use_sma_macd):
    return {
        s: indicators.vectorized_signals(
            dfs[s], use_triple_st=use_triple_st, use_sma_macd=use_sma_macd
        ).values
        for s in dfs
    }

def bah_period(symbols, tf, years, old_period=False):
    n   = ms.CANDLES_PER_YEAR[tf] * years
    n4y = ms.CANDLES_PER_YEAR[tf] * 4
    rets = []
    for sym in symbols:
        df = ms.get_df_8y(sym, tf) if old_period else ms.get_df(sym, tf)
        if df is None or len(df) < 10:
            continue
        if old_period:
            if len(df) <= n4y:
                continue
            dp = df.iloc[:-n4y].tail(n)
            if len(dp) < int(n * 0.9):
                continue
        else:
            dp = df.tail(n)
        s, e = dp["close"].iloc[0], dp["close"].iloc[-1]
        if s > 0:
            rets.append((e - s) / s * 100)
    return round(sum(rets) / len(rets), 1) if rets else None

def bah_year(symbols, tf, year):
    rets = []
    for sym in symbols:
        df = ms.get_df_for_year(sym, tf, year)
        if df is None or len(df) < 10:
            continue
        s, e = df["close"].iloc[0], df["close"].iloc[-1]
        if s > 0:
            rets.append((e - s) / s * 100)
    return round(sum(rets) / len(rets), 1) if rets else None

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    N_WORKERS = os.cpu_count() or 4

    n_recent = len(PORTFOLIOS)*2*2*len(ALL_TF)*len(PERIODS)*len(STRATEGIES)*len(CAPITAL_SCHEMES)*len(TP_MODES)
    n_old    = len(PORTFOLIOS)*2*2*len(OLD_TF)*len(PERIODS)*len(STRATEGIES)*len(CAPITAL_SCHEMES)*len(TP_MODES)
    n_year   = len(CALENDAR_YEARS)*len(PORTFOLIOS)*2*2*len(ALL_TF)*len(STRATEGIES)*len(CAPITAL_SCHEMES)*len(TP_MODES)

    print(f"\n{Fore.CYAN}{'='*110}")
    print("  TP RANKING FULL — toutes dimensions + 8 modes TP | multiprocessing fork")
    print(f"  ~{n_recent+n_old+n_year:,} tâches | {N_WORKERS} workers")
    print(f"{'='*110}{Style.RESET_ALL}")

    # ---- Chargement des données ------------------------------------------------
    all_symbols = list(dict.fromkeys(s for syms in PORTFOLIOS.values() for s in syms))

    print(f"\n{Fore.CYAN}Données récentes...{Style.RESET_ALL}", flush=True)
    data_cache.prefetch_all(all_symbols, ALL_TF, verbose=False)
    for sym in all_symbols:
        for tf in ALL_TF:
            ms.get_df(sym, tf)

    print(f"{Fore.CYAN}Données anciennes...{Style.RESET_ALL}", flush=True)
    data_cache.prefetch_all_8y(all_symbols, OLD_TF, verbose=False)
    for sym in all_symbols:
        for tf in OLD_TF:
            ms.get_df_8y(sym, tf)

    print(f"{Fore.CYAN}Fear & Greed...{Style.RESET_ALL}", flush=True)
    _FG_DATA.update(fear_greed.load(verbose=False))

    # ---- Buy-and-hold ----------------------------------------------------------
    print(f"{Fore.CYAN}Buy-and-hold...{Style.RESET_ALL}", flush=True)
    for port, syms in PORTFOLIOS.items():
        for tf in ALL_TF:
            for years in PERIODS:
                _BAH_CACHE[(port, tf, years, False)] = bah_period(syms, tf, years, False)
        for tf in OLD_TF:
            for years in PERIODS:
                _BAH_CACHE[(port, tf, years, True)]  = bah_period(syms, tf, years, True)
        for year in CALENDAR_YEARS:
            for tf in ALL_TF:
                _BAH_CACHE[(port, tf, year)] = bah_year(syms, tf, year)

    # ---- Pré-calcul des signaux ------------------------------------------------
    total_sig = (len(PORTFOLIOS)*len(ALL_TF)*len(PERIODS)*len(STRATEGIES)
                 + len(PORTFOLIOS)*len(OLD_TF)*len(PERIODS)*len(STRATEGIES)
                 + len(CALENDAR_YEARS)*len(PORTFOLIOS)*len(ALL_TF)*len(STRATEGIES))

    print(f"{Fore.CYAN}Pré-calcul des signaux ({total_sig:,} combos)...{Style.RESET_ALL}", flush=True)
    done = 0
    for port, syms in PORTFOLIOS.items():
        for tf in ALL_TF:
            for years in PERIODS:
                dfs = _slice_recent(syms, tf, years)
                for sn, ust, usm, _ in STRATEGIES:
                    key = ("rec", port, tf, years, sn)
                    _SIG_CACHE[key] = (dfs, _compute_sigs(dfs, ust, usm)) if dfs else ({}, {})
                    done += 1
        for tf in OLD_TF:
            for years in PERIODS:
                dfs = _slice_old(syms, tf, years)
                for sn, ust, usm, _ in STRATEGIES:
                    key = ("old", port, tf, years, sn)
                    _SIG_CACHE[key] = (dfs, _compute_sigs(dfs, ust, usm)) if dfs else ({}, {})
                    done += 1
        for year in CALENDAR_YEARS:
            for tf in ALL_TF:
                dfs = _slice_year(syms, tf, year)
                for sn, ust, usm, _ in STRATEGIES:
                    key = ("year", port, tf, year, sn)
                    _SIG_CACHE[key] = (dfs, _compute_sigs(dfs, ust, usm)) if dfs else ({}, {})
                    done += 1
        print(f"\r  {done}/{total_sig}", end="", flush=True)
    print(f"\n  {done} combos prêts")

    # ---- Construction de la liste de tâches ------------------------------------
    print(f"{Fore.CYAN}Construction des tâches...{Style.RESET_ALL}", flush=True)
    tasks_recent, tasks_old, tasks_year = [], [], []

    for port, syms in PORTFOLIOS.items():
        pl = PORT_LABELS[port]
        for mode in ("single", "multi"):
            ml = "1pos" if mode == "single" else "multi"
            for use_sl in (True, False):
                sl = "avecSL" if use_sl else "sansSL"
                for tf in ALL_TF:
                    for years, period_label in PERIODS.items():
                        bah = _BAH_CACHE.get((port, tf, years, False))
                        for sn, _, _, ms_score in STRATEGIES:
                            for cl, pos_pct, max_tr in CAPITAL_SCHEMES:
                                for tl, ut, tp, at in TP_MODES:
                                    tasks_recent.append((
                                        ("rec", port, tf, years, sn),
                                        pl, ml, sl, tf, period_label,
                                        sn, ms_score, cl, pos_pct, max_tr,
                                        tl, ut, tp, at, bah, False, False
                                    ))
                for tf in OLD_TF:
                    for years, period_label in PERIODS.items():
                        bah_old = _BAH_CACHE.get((port, tf, years, True))
                        for sn, _, _, ms_score in STRATEGIES:
                            for cl, pos_pct, max_tr in CAPITAL_SCHEMES:
                                for tl, ut, tp, at in TP_MODES:
                                    tasks_old.append((
                                        ("old", port, tf, years, sn),
                                        pl, ml, sl, tf, period_label,
                                        sn, ms_score, cl, pos_pct, max_tr,
                                        tl, ut, tp, at, bah_old, True, False
                                    ))
                for year in CALENDAR_YEARS:
                    for tf in ALL_TF:
                        bah = _BAH_CACHE.get((port, tf, year))
                        for sn, _, _, ms_score in STRATEGIES:
                            for cl, pos_pct, max_tr in CAPITAL_SCHEMES:
                                for tl, ut, tp, at in TP_MODES:
                                    tasks_year.append((
                                        ("year", port, tf, year, sn),
                                        pl, ml, sl, tf, str(year),
                                        sn, ms_score, cl, pos_pct, max_tr,
                                        tl, ut, tp, at, bah, False, True
                                    ))

    all_tasks = tasks_recent + tasks_old + tasks_year
    print(f"  {len(all_tasks):,} tâches créées")

    # ---- Exécution parallèle (fork) --------------------------------------------
    ctx = multiprocessing.get_context("fork")
    all_rows = []

    for label, tasks in [
        (f"Simulations récentes ({len(tasks_recent):,})", tasks_recent),
        (f"Simulations anciennes ({len(tasks_old):,})",   tasks_old),
        (f"Simulations annuelles ({len(tasks_year):,})",  tasks_year),
    ]:
        print(f"\n{Fore.CYAN}{label}...{Style.RESET_ALL}", flush=True)
        done = 0
        total = len(tasks)
        with ctx.Pool(processes=N_WORKERS) as pool:
            for result in pool.imap_unordered(_worker, tasks, chunksize=200):
                if result is not None:
                    all_rows.append(result)
                done += 1
                if done % 5000 == 0 or done == total:
                    print(f"\r  {done:,}/{total:,}", end="", flush=True)
        print()

    # ---- Nettoyage des colonnes internes ---------------------------------------
    df = pd.DataFrame(all_rows)
    df = df.drop(columns=["is_old", "sig_key"], errors="ignore")
    df = df[df["return_%"].notna()].copy()
    df = df.sort_values("return_%", ascending=False).reset_index(drop=True)
    df.index += 1
    print(f"\n{Fore.GREEN}  {len(df):,} simulations valides{Style.RESET_ALL}")

    # ---- Résumé par TP ---------------------------------------------------------
    def color(v, fmt="+.1f"):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return "—"
        c = Fore.GREEN if v > 0 else (Fore.RED if v < 0 else "")
        return f"{c}{v:{fmt}}%{Style.RESET_ALL}"

    print(f"\n{Fore.YELLOW}{'='*90}")
    print("  MÉDIANE return% par TP (toutes dimensions)")
    print(f"{'='*90}{Style.RESET_ALL}")
    df_rec = df[~df["période"].str.match(r"^\d{4}$", na=False)]
    for src_label, df_src in [("Périodes récentes 1-4ans", df_rec), ("Toutes périodes", df)]:
        med = df_src.groupby("tp")["return_%"].median().reindex([m[0] for m in TP_MODES])
        print(f"\n  {src_label}:")
        for tp_lbl, val in med.items():
            if math.isnan(val):
                continue
            bar = "█" * max(0, int(val / 3)) if val > 0 else "░" * max(0, int(-val / 3))
            c = Fore.GREEN if val > 0 else Fore.RED
            print(f"    {tp_lbl:<8} {c}{val:+6.1f}%{Style.RESET_ALL}  {bar}")

    # ---- TOP 50 ----------------------------------------------------------------
    print(f"\n{Fore.YELLOW}{'='*160}")
    print("  TOP 50 — meilleures simulations")
    print(f"{'='*160}{Style.RESET_ALL}")
    headers = ["#","TP","Portfolio","Mode","SL","TF","Période","Stratégie","Capital",
               "Return%","DD%","Win%","Trades","PF","B&H%","Alpha"]
    top_rows = []
    for i in range(min(50, len(df))):
        row = df.iloc[i]
        top_rows.append([
            i+1, row["tp"], row["portfolio"], row["mode"], row["sl"],
            row["timeframe"], row["période"], row["stratégie"], row["capital"],
            color(row["return_%"]), color(row["drawdown_%"]),
            f"{row['win_%']:.0f}%" if row["win_%"] is not None else "—",
            f"{row['trades']:.0f}" if row["trades"] is not None else "—",
            f"{row['pf']:.2f}" if row["pf"] is not None else "—",
            f"{row['bah_%']:+.1f}%" if row["bah_%"] is not None else "—",
            color(row["alpha_%"]),
        ])
    print(tabulate(top_rows, headers=headers, tablefmt="rounded_outline"))

    # ---- Régularité ------------------------------------------------------------
    year_df = df[df["période"].str.match(r"^\d{4}$", na=False)].copy()
    reg_df  = pd.DataFrame()
    if not year_df.empty:
        config_cols = ["tp","portfolio","mode","sl","timeframe","stratégie","capital"]
        grp = year_df.groupby(config_cols)["return_%"]
        reg_df = pd.DataFrame({
            "n_années":  grp.count(),
            "moy_%":     grp.mean().round(1),
            "min_%":     grp.min().round(1),
            "max_%":     grp.max().round(1),
            "std_%":     grp.std().round(1),
            "n_positif": grp.apply(lambda x: int((x > 0).sum())),
        }).reset_index()
        reg_df["pct_positif"] = (reg_df["n_positif"] / reg_df["n_années"] * 100).round(0).astype(int)
        reg_df["score"]       = (reg_df["moy_%"] * (reg_df["pct_positif"] / 100) ** 2).round(1)
        reg_df = reg_df[reg_df["n_années"] >= 4].sort_values("score", ascending=False).reset_index(drop=True)
        reg_df.index += 1

        print(f"\n{Fore.CYAN}{'='*130}")
        print("  TOP 30 RÉGULARITÉ  (score = moy% × pct_positif²)")
        print(f"{'='*130}{Style.RESET_ALL}")
        reg_rows = []
        for i in range(min(30, len(reg_df))):
            row = reg_df.iloc[i]
            reg_rows.append([
                i+1, row["tp"], row["portfolio"], row["mode"], row["sl"],
                row["timeframe"], row["stratégie"], row["capital"],
                row["n_années"],
                f"{Fore.GREEN if row['moy_%']>=0 else Fore.RED}{row['moy_%']:+.1f}%{Style.RESET_ALL}",
                f"{Fore.RED if row['min_%']<0 else Fore.GREEN}{row['min_%']:+.1f}%{Style.RESET_ALL}",
                f"{row['max_%']:+.1f}%", f"{row['std_%']:.1f}",
                f"{row['pct_positif']}%",
                f"{Fore.CYAN}{row['score']:+.1f}{Style.RESET_ALL}",
            ])
        print(tabulate(reg_rows,
                       headers=["#","TP","Portfolio","Mode","SL","TF","Stratégie","Capital",
                                 "N_ans","Moy%","Min%","Max%","Std","Pct+","Score"],
                       tablefmt="rounded_outline"))

        print(f"\n{Fore.CYAN}  Score régularité médian par TP :{Style.RESET_ALL}")
        score_med = reg_df.groupby("tp")["score"].median().reindex([m[0] for m in TP_MODES])
        for tp_lbl, val in score_med.items():
            if math.isnan(val):
                continue
            bar = "█" * max(0, int(val / 2)) if val > 0 else "░" * max(0, int(-val / 2))
            c = Fore.GREEN if val > 0 else Fore.RED
            print(f"    {tp_lbl:<8} {c}{val:+6.1f}{Style.RESET_ALL}  {bar}")

    # ---- Export ----------------------------------------------------------------
    csv_path  = "tp_ranking_full_results.csv"
    xlsx_path = "tp_ranking_full_results.xlsx"

    df.to_csv(csv_path, index=True, float_format="%.2f")
    print(f"\n{Fore.GREEN}Export → {csv_path} ({len(df):,} lignes){Style.RESET_ALL}")

    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GH = PatternFill("solid", fgColor="C6EFCE"); GC = PatternFill("solid", fgColor="EBF5EB")
        RH = PatternFill("solid", fgColor="FFC7CE"); RC = PatternFill("solid", fgColor="FFE7EA")
        GRH = PatternFill("solid", fgColor="D9D9D9")
        BOLD = Font(bold=True); CENTER = Alignment(horizontal="center")
        thin = Side(style="thin", color="BBBBBB")
        BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style(ws, df_s, green, red=None):
            red = red or set()
            for ci, cn in enumerate(df_s.columns, 1):
                ml = max(len(str(cn)), df_s[cn].astype(str).str.len().max())
                ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 22)
            for ci, cn in enumerate(df_s.columns, 1):
                hf = GH if cn in green else (RH if cn in red else GRH)
                cf = GC if cn in green else (RC if cn in red else None)
                hc = ws.cell(row=1, column=ci)
                hc.fill = hf; hc.font = BOLD; hc.alignment = CENTER; hc.border = BRD
                for ri in range(2, len(df_s) + 2):
                    cell = ws.cell(row=ri, column=ci)
                    if cf: cell.fill = cf
                    cell.alignment = CENTER; cell.border = BRD
            ws.freeze_panes = "A2"

        GREEN = {"return_%","drawdown_%","win_%","trades","pf","bah_%","alpha_%","capital","tp","volume_annuel_%"}

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_xl = df.reset_index().rename(columns={"index": "rang"})
            df_xl.to_excel(writer, index=False, sheet_name="Ranking")
            style(writer.sheets["Ranking"], df_xl, GREEN)

            if not reg_df.empty:
                rx = reg_df.reset_index().rename(columns={"index": "rang"})
                rx.to_excel(writer, index=False, sheet_name="Régularité")
                style(writer.sheets["Régularité"], rx,
                      green={"moy_%","score","pct_positif","tp"}, red={"min_%"})

            grp_tp = df.groupby("tp")["return_%"].agg(
                médiane="median", moyenne="mean", max="max", min="min"
            ).round(2).reindex([m[0] for m in TP_MODES]).reset_index()
            grp_tp.to_excel(writer, index=False, sheet_name="Résumé TP")
            style(writer.sheets["Résumé TP"], grp_tp,
                  green={"médiane","moyenne","max","tp"}, red={"min"})

        print(f"{Fore.GREEN}Export → {xlsx_path} (3 feuilles){Style.RESET_ALL}")
    except ImportError:
        print(f"{Fore.YELLOW}openpyxl non installé{Style.RESET_ALL}")
