"""
Étape 19 — Capital pondéré par régularité dans full_ranking_results

Ajoute au tableau existant (full_ranking_results.csv / .xlsx) les simulations
utilisant sim_multi_weighted() : chaque symbole reçoit un % de capital
proportionnel à sa régularité historique (score = moy_% × pct_positif²).

Deux variantes :
  "pondéré"        → floor=0.02 (2% minimum par symbole)
  "pondéré-strict" → floor=0.0  (tout le poids sur les meilleurs)

Portfolios concernés : Top 20, Top 10, Top 5 (seuls avec ≥ 3 symboles)
Mode : multi uniquement
"""

import pandas as pd
import numpy as np
from colorama import Fore, Style, init
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import config
import data_cache
import fear_greed
import multi_sim as ms
from multi_sim import sim_multi_weighted

init(autoreset=True)

# ---------------------------------------------------------------------------
# Dimensions
# ---------------------------------------------------------------------------
ALL_TF  = ["30m", "1h", "2h", "4h", "6h", "12h", "1d"]
PERIODS = {1: "1an", 2: "2ans", 3: "3ans", 4: "4ans"}
CALENDAR_YEARS = list(range(2016, 2027))

PORTFOLIOS = {
    "top20": config.SYMBOLS,
    "top10": ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT",
              "ADA/USDT","AVAX/USDT","DOT/USDT","LINK/USDT","MATIC/USDT"],
    "top5":  ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT"],
}
PORT_LABELS = {"top20": "Top 20", "top10": "Top 10", "top5": "Top 5"}

STRATEGIES = [
    ("baseline",   False, True,  3),
    ("+TripleST",  True,  True,  3),
    ("épurée",     False, False, 3),
    ("épurée+ST",  True,  False, 3),
]

WEIGHT_VARIANTS = [
    ("pondéré",        0.02),
    ("pondéré-strict", 0.00),
]

SCORE_YEARS = list(range(2018, 2025))  # 2018-2024
MIN_YEARS   = 4   # pour les TF avec 7 ans dispo (2h+)
MIN_YEARS_SHORT = 2  # pour 30m/1h qui n'ont que 3 ans calendaires


# ---------------------------------------------------------------------------
# Calcul des poids de régularité
# ---------------------------------------------------------------------------
def scores_to_weights(scores: dict, symbols: list, floor_pct: float = 0.0) -> dict:
    """Convertit un dict {sym: score} en poids {sym: fraction_capital}."""
    scores_pos = {s: max(scores.get(s, 0.0), 0.0) for s in symbols}
    total = sum(scores_pos.values())

    if total == 0:
        return {s: 1.0 / len(symbols) for s in symbols}

    poids_raw = {s: scores_pos[s] / total for s in symbols}

    if floor_pct == 0.0:
        return poids_raw

    n = len(symbols)
    total_floor = floor_pct * n
    if total_floor >= 1.0:
        return {s: 1.0 / n for s in symbols}

    remaining = 1.0 - total_floor
    return {s: floor_pct + poids_raw[s] * remaining for s in symbols}


def compute_weights(symbols, tf, use_triple_st, use_sma_macd, fg, floor_pct):
    """
    Simule chaque symbole isolé sur 2018-2024, calcule le score de régularité
    et retourne les poids finaux {sym: fraction_capital}.
    """
    import data_cache as _dc
    # 30m/1h n'ont que 3 années calendaires dispo → seuil abaissé
    min_y = MIN_YEARS_SHORT if tf not in _dc.TF_WITH_8Y else MIN_YEARS

    scores = {}
    for sym in symbols:
        rets = []
        for year in SCORE_YEARS:
            df_y = ms.get_df_for_year(sym, tf, year)
            if df_y is None or len(df_y) < 10:
                continue
            r = ms.sim_multi_on_dfs(
                {sym: df_y}, use_sl=False, fg=fg,
                use_triple_st=use_triple_st, use_sma_macd=use_sma_macd, tf=tf
            )
            ret = r.get("return_%")
            if ret is not None:
                rets.append(ret)
        if len(rets) >= min_y:
            moy = sum(rets) / len(rets)
            pct_pos = sum(1 for v in rets if v > 0) / len(rets)
            scores[sym] = moy * (pct_pos ** 2)
        else:
            scores[sym] = 0.0

    return scores_to_weights(scores, symbols, floor_pct)


# ---------------------------------------------------------------------------
# BAH helpers (reprises de full_ranking.py)
# ---------------------------------------------------------------------------
def buy_and_hold(symbols, tf, years):
    n = ms.CANDLES_PER_YEAR[tf] * years
    returns = []
    for sym in symbols:
        df = ms.get_df(sym, tf)
        if df is None or len(df) < 10:
            continue
        dp = df.tail(n)
        s, e = dp["close"].iloc[0], dp["close"].iloc[-1]
        if s > 0:
            returns.append((e - s) / s * 100)
    return round(sum(returns) / len(returns), 1) if returns else None


def buy_and_hold_year(symbols, tf, year):
    returns = []
    for sym in symbols:
        df = ms.get_df_for_year(sym, tf, year)
        if df is None or len(df) < 10:
            continue
        s, e = df["close"].iloc[0], df["close"].iloc[-1]
        if s > 0:
            returns.append((e - s) / s * 100)
    return round(sum(returns) / len(returns), 1) if returns else None


# ---------------------------------------------------------------------------
# Simulation pondérée sur une période glissante
# ---------------------------------------------------------------------------
def run_weighted_period(symbols, tf, years, weights, use_sl, use_triple_st, use_sma_macd, fg):
    n = ms.CANDLES_PER_YEAR[tf] * years
    dfs = {}
    for sym in symbols:
        df = ms.get_df(sym, tf)
        if df is not None and len(df) >= 10:
            sliced = df.tail(n).reset_index(drop=True)
            if len(sliced) >= int(n * 0.5):
                dfs[sym] = sliced
    if not dfs:
        return {}
    original = config.MIN_SCORE_TO_TRADE
    config.MIN_SCORE_TO_TRADE = 3
    try:
        return sim_multi_weighted(
            dfs, pos_pct_per_symbol=weights, fg=fg,
            use_triple_st=use_triple_st, use_sma_macd=use_sma_macd,
            use_sl=use_sl, tf=tf
        )
    finally:
        config.MIN_SCORE_TO_TRADE = original


def run_weighted_year(symbols, tf, year, weights, use_sl, use_triple_st, use_sma_macd, fg):
    dfs = {}
    for sym in symbols:
        df = ms.get_df_for_year(sym, tf, year)
        if df is not None:
            dfs[sym] = df
    if not dfs:
        return {}
    original = config.MIN_SCORE_TO_TRADE
    config.MIN_SCORE_TO_TRADE = 3
    try:
        return sim_multi_weighted(
            dfs, pos_pct_per_symbol=weights, fg=fg,
            use_triple_st=use_triple_st, use_sma_macd=use_sma_macd,
            use_sl=use_sl, tf=tf
        )
    finally:
        config.MIN_SCORE_TO_TRADE = original


# ---------------------------------------------------------------------------
# Export Excel (même style que full_ranking)
# ---------------------------------------------------------------------------
def export_xlsx(df_combined, path="full_ranking_results.xlsx"):
    GREEN_H  = PatternFill("solid", fgColor="C6EFCE")
    BLUE_H   = PatternFill("solid", fgColor="BDD7EE")
    GREY_H   = PatternFill("solid", fgColor="D9D9D9")
    GOLD_H   = PatternFill("solid", fgColor="FFD700")
    RED_FILL = PatternFill("solid", fgColor="FFCCCC")
    GRN_FILL = PatternFill("solid", fgColor="E2EFDA")
    thin     = Side(style="thin", color="CCCCCC")
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
    BOLD     = Font(bold=True)
    CENTER   = Alignment(horizontal="center")

    col_widths = {
        "portfolio": 10, "mode": 7, "sl": 8, "timeframe": 10,
        "période": 8, "stratégie": 12, "capital": 16,
        "return_%": 10, "drawdown_%": 11, "win_%": 8, "trades": 7,
        "pf": 6, "dur.moy_j": 10, "non_ferm.": 10,
        "bah_%": 8, "alpha_%": 9, "ret_old_%": 10, "dd_old_%": 9,
        "bah_old_%": 10, "alpha_old_%": 11, "delta_%": 8, "classement": 11,
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_combined.to_excel(writer, index=False, sheet_name="Ranking")
        ws = writer.sheets["Ranking"]

        cols = list(df_combined.columns)
        cls_col   = cols.index("classement") + 1 if "classement" in cols else None
        ret_col   = cols.index("return_%") + 1
        alpha_col = cols.index("alpha_%") + 1

        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = BOLD
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 10)
            if col_name == "classement":
                cell.fill = GOLD_H
            elif any(x in col_name for x in ["alpha", "bah"]):
                cell.fill = GREEN_H
            elif "old" in col_name:
                cell.fill = BLUE_H
            else:
                cell.fill = GREY_H

        for row_idx in range(2, len(df_combined) + 2):
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = CENTER
                ws.cell(row=row_idx, column=col_idx).border = BORDER

            if cls_col:
                cls_val = ws.cell(row=row_idx, column=cls_col).value
                if cls_val is not None and not (isinstance(cls_val, float) and np.isnan(cls_val)):
                    ws.cell(row=row_idx, column=cls_col).fill = GOLD_H
                    ws.cell(row=row_idx, column=cls_col).font = Font(bold=True)

            ret_val = ws.cell(row=row_idx, column=ret_col).value
            if isinstance(ret_val, (int, float)):
                ws.cell(row=row_idx, column=ret_col).fill = GRN_FILL if ret_val > 0 else RED_FILL

            alpha_val = ws.cell(row=row_idx, column=alpha_col).value
            if isinstance(alpha_val, (int, float)):
                ws.cell(row=row_idx, column=alpha_col).fill = GRN_FILL if alpha_val > 0 else RED_FILL

        ws.freeze_panes = "A2"

    print(f"  Ranking : OK ({len(df_combined)} lignes)")

    # Régularité sheet
    config_cols = ["portfolio", "sl", "timeframe", "stratégie", "capital"]
    df_multi = df_combined[df_combined["mode"] == "multi"].copy()
    years_list = [str(y) for y in range(2016, 2027)]
    df_years = df_multi[df_multi["période"].isin(years_list)].copy()

    reg = df_years.groupby(config_cols).apply(
        lambda g: pd.Series({
            "n_années": len(g),
            "pct_positif": round((g["return_%"] > 0).mean() * 100, 1),
            "moy_%": round(g["return_%"].mean(), 2),
            "score_reg": round(g["return_%"].mean() * ((g["return_%"] > 0).mean() ** 2), 2),
        })
    ).reset_index()

    df_4ans = df_multi[df_multi["période"] == "4ans"].copy()
    g4 = df_4ans.groupby(config_cols).agg(
        return_4ans=("return_%", "mean"),
        alpha_4ans=("alpha_%", "mean"),
        dd_4ans=("drawdown_%", "mean"),
    ).reset_index()

    reg = reg.merge(g4, on=config_cols, how="left")
    reg = reg[reg["n_années"] >= 4].sort_values("score_reg", ascending=False).reset_index(drop=True)
    reg.insert(0, "rang_reg", range(1, len(reg) + 1))

    wb = load_workbook(path)
    if "Régularité" in wb.sheetnames:
        del wb["Régularité"]
    ws2 = wb.create_sheet("Régularité")

    reg_cols = list(reg.columns)
    for col_idx, col_name in enumerate(reg_cols, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.fill = GREY_H
        cell.alignment = CENTER
        cell.border = BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 2, 10)

    for row_idx, (_, row) in enumerate(reg.iterrows(), start=2):
        for col_idx, col_name in enumerate(reg_cols, start=1):
            val = row[col_name]
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CENTER
            cell.border = BORDER
            if col_name in ("score_reg", "pct_positif", "moy_%", "alpha_4ans", "return_4ans"):
                cell.fill = GRN_FILL if (val or 0) > 0 else RED_FILL

    ws2.freeze_panes = "A2"
    wb.save(path)
    print(f"  Régularité : OK ({len(reg)} lignes)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{Fore.CYAN}{'='*90}")
    print("  ÉTAPE 19 — Capital pondéré par régularité")
    print(f"{'='*90}{Style.RESET_ALL}")

    all_symbols = list(dict.fromkeys(s for syms in PORTFOLIOS.values() for s in syms))

    # ---- Chargement données ------------------------------------------------
    print(f"\n{Fore.CYAN}Chargement données récentes (4 ans)...{Style.RESET_ALL}")
    data_cache.prefetch_all(all_symbols, ALL_TF, verbose=False)
    for sym in all_symbols:
        for tf in ALL_TF:
            ms.get_df(sym, tf)

    print(f"{Fore.CYAN}Chargement données 10 ans (années calendaires)...{Style.RESET_ALL}")
    data_cache.prefetch_all_10y(all_symbols, data_cache.TF_WITH_10Y, verbose=False)
    for sym in all_symbols:
        for tf in data_cache.TF_WITH_10Y:
            ms.get_df_10y(sym, tf)

    fg_data = fear_greed.load(verbose=False)
    print(f"  Fear & Greed : {len(fg_data)} jours")

    # ---- BAH par période et par année --------------------------------------
    print(f"{Fore.CYAN}Calcul Buy-and-Hold...{Style.RESET_ALL}")
    bah_cache      = {}
    bah_year_cache = {}
    for port, syms in PORTFOLIOS.items():
        for tf in ALL_TF:
            for years in PERIODS:
                bah_cache[(port, tf, years)] = buy_and_hold(syms, tf, years)
            for year in CALENDAR_YEARS:
                bah_year_cache[(port, tf, year)] = buy_and_hold_year(syms, tf, year)

    # ---- Calcul des poids (une fois par portfolio × stratégie × TF × variant) ---
    print(f"\n{Fore.CYAN}Calcul des poids de régularité...{Style.RESET_ALL}")
    weights_cache = {}
    total_w = len(PORTFOLIOS) * len(STRATEGIES) * len(ALL_TF) * len(WEIGHT_VARIANTS)
    done_w  = 0
    for port, syms in PORTFOLIOS.items():
        for strat_name, use_ts, use_sm, _ in STRATEGIES:
            for tf in ALL_TF:
                for cap_label, floor_pct in WEIGHT_VARIANTS:
                    key = (port, strat_name, tf, floor_pct)
                    weights_cache[key] = compute_weights(
                        syms, tf, use_ts, use_sm, fg_data, floor_pct
                    )
                    done_w += 1
                    print(f"\r  {done_w}/{total_w} poids calculés", end="", flush=True)
    print()

    # ---- Simulations -------------------------------------------------------
    new_rows = []

    # Périodes glissantes (1an, 2ans, 3ans, 4ans)
    total_p = (len(PORTFOLIOS) * 2 * len(ALL_TF) * len(PERIODS)
               * len(STRATEGIES) * len(WEIGHT_VARIANTS))
    done_p  = 0
    print(f"{Fore.CYAN}Simulations périodes ({total_p})...{Style.RESET_ALL}")

    for port, syms in PORTFOLIOS.items():
        for use_sl in (True, False):
            sl_label = "avecSL" if use_sl else "sansSL"
            for tf in ALL_TF:
                for years, période_label in PERIODS.items():
                    bah = bah_cache.get((port, tf, years))
                    for strat_name, use_ts, use_sm, _ in STRATEGIES:
                        for cap_label, floor_pct in WEIGHT_VARIANTS:
                            w_key = (port, strat_name, tf, floor_pct)
                            weights = weights_cache[w_key]
                            try:
                                r = run_weighted_period(
                                    syms, tf, years, weights, use_sl, use_ts, use_sm, fg_data
                                )
                            except Exception:
                                r = {}
                            ret = r.get("return_%")
                            new_rows.append({
                                "portfolio":   PORT_LABELS[port],
                                "mode":        "multi",
                                "sl":          sl_label,
                                "timeframe":   tf,
                                "période":     période_label,
                                "stratégie":   strat_name,
                                "capital":     cap_label,
                                "return_%":    ret,
                                "drawdown_%":  r.get("drawdown_%"),
                                "win_%":       r.get("win_%"),
                                "trades":      r.get("trades"),
                                "pf":          r.get("profit_factor"),
                                "dur.moy_j":   r.get("avg_duration_j"),
                                "non_ferm.":   r.get("unclosed"),
                                "bah_%":       bah,
                                "alpha_%":     round(ret - bah, 1) if ret is not None and bah is not None else None,
                                "ret_old_%":   None,
                                "dd_old_%":    None,
                                "bah_old_%":   None,
                                "alpha_old_%": None,
                                "delta_%":     None,
                                "classement":  pd.NA,
                            })
                            done_p += 1
                print(f"\r  {done_p}/{total_p}", end="", flush=True)
    print()

    # Années calendaires
    total_y = (len(PORTFOLIOS) * 2 * len(ALL_TF) * len(CALENDAR_YEARS)
               * len(STRATEGIES) * len(WEIGHT_VARIANTS))
    done_y  = 0
    print(f"{Fore.CYAN}Simulations années calendaires ({total_y})...{Style.RESET_ALL}")

    for port, syms in PORTFOLIOS.items():
        for use_sl in (True, False):
            sl_label = "avecSL" if use_sl else "sansSL"
            for tf in ALL_TF:
                for year in CALENDAR_YEARS:
                    bah_y = bah_year_cache.get((port, tf, year))
                    for strat_name, use_ts, use_sm, _ in STRATEGIES:
                        for cap_label, floor_pct in WEIGHT_VARIANTS:
                            w_key = (port, strat_name, tf, floor_pct)
                            weights = weights_cache[w_key]
                            try:
                                r = run_weighted_year(
                                    syms, tf, year, weights, use_sl, use_ts, use_sm, fg_data
                                )
                            except Exception:
                                r = {}
                            ret = r.get("return_%")
                            new_rows.append({
                                "portfolio":   PORT_LABELS[port],
                                "mode":        "multi",
                                "sl":          sl_label,
                                "timeframe":   tf,
                                "période":     str(year),
                                "stratégie":   strat_name,
                                "capital":     cap_label,
                                "return_%":    ret,
                                "drawdown_%":  r.get("drawdown_%"),
                                "win_%":       r.get("win_%"),
                                "trades":      r.get("trades"),
                                "pf":          r.get("profit_factor"),
                                "dur.moy_j":   r.get("avg_duration_j"),
                                "non_ferm.":   r.get("unclosed"),
                                "bah_%":       bah_y,
                                "alpha_%":     round(ret - bah_y, 1) if ret is not None and bah_y is not None else None,
                                "ret_old_%":   None,
                                "dd_old_%":    None,
                                "bah_old_%":   None,
                                "alpha_old_%": None,
                                "delta_%":     None,
                                "classement":  pd.NA,
                            })
                            done_y += 1
                print(f"\r  {done_y}/{total_y}", end="", flush=True)
    print()

    # ---- Merge avec données existantes ------------------------------------
    print(f"\n{Fore.CYAN}Merge avec full_ranking_results.csv...{Style.RESET_ALL}")
    df_existing = pd.read_csv("full_ranking_results.csv", index_col=0)
    df_new      = pd.DataFrame(new_rows)

    # Assurer que classement est Int64 nullable
    if "classement" not in df_existing.columns:
        df_existing["classement"] = pd.NA
    df_existing["classement"] = df_existing["classement"].astype(pd.Int64Dtype())
    df_new["classement"]      = df_new["classement"].astype(pd.Int64Dtype())

    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    df_combined.sort_values("return_%", ascending=False, inplace=True, na_position="last")
    df_combined.reset_index(drop=True, inplace=True)

    df_combined.to_csv("full_ranking_results.csv")
    print(f"  CSV sauvegardé : {len(df_combined)} lignes au total (+{len(df_new)} nouvelles)")

    # ---- Aperçu des meilleurs pondérés -------------------------------------
    print(f"\n{Fore.YELLOW}{'='*90}")
    print("  APERÇU — Meilleures configs pondérées (période 4ans, multi, alpha > 0)")
    print(f"{'='*90}{Style.RESET_ALL}")

    df_pond = df_new[
        (df_new["mode"] == "multi") &
        (df_new["période"] == "4ans") &
        (df_new["alpha_%"].notna()) &
        (df_new["alpha_%"] > 0)
    ].sort_values("alpha_%", ascending=False).head(15)

    from tabulate import tabulate
    cols_show = ["portfolio","stratégie","capital","timeframe","sl","return_%","bah_%","alpha_%","drawdown_%","win_%"]
    print(tabulate(df_pond[cols_show].values.tolist(), headers=cols_show, tablefmt="rounded_outline", floatfmt=".1f"))

    # ---- Export Excel -------------------------------------------------------
    print(f"\n{Fore.CYAN}Export Excel...{Style.RESET_ALL}")
    export_xlsx(df_combined)
    print(f"\n{Fore.GREEN}✓ full_ranking_results.xlsx mis à jour{Style.RESET_ALL}")
    print()
