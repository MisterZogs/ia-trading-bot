"""
Comparaison des modes de Take-Profit sur toutes les stratégies.

Dimensions :
  TP       : 2%, 3%, 5%, 10%, 15%, 20%, ATR dynamique, sans TP
  TF       : 30m, 1h, 2h, 4h, 6h, 12h, 1d
  Stratégie: baseline, +TripleST, épurée, épurée+ST
  SL       : avecSL, sansSL
  Portfolio: Top 20

Périodes :
  - Années calendaires 2018–2025 (pour score de régularité)
  - 4 ans récents (résumé global)

Usage :
    python3 tp_ranking.py
"""

import math
import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

import config
import data_cache
import multi_sim as ms
import fear_greed
import indicators

init(autoreset=True)

# ---------------------------------------------------------------------------
# Dimensions
# ---------------------------------------------------------------------------
ALL_TF     = ["30m", "1h", "2h", "4h", "6h", "12h", "1d"]
SYMBOLS    = config.SYMBOLS   # Top 20
YEARS      = list(range(2018, 2026))   # 2018 → 2025

STRATEGIES = [
    ("baseline",  False, True),
    ("+TripleST", True,  True),
    ("épurée",    False, False),
    ("épurée+ST", True,  False),
]

# (label, use_tp, tp_pct, atr_tp)
TP_MODES = [
    ("2%",    True,  0.02,  False),
    ("3%",    True,  0.03,  False),
    ("5%",    True,  0.05,  False),
    ("10%",   True,  0.10,  False),
    ("15%",   True,  0.15,  False),
    ("20%",   True,  0.20,  False),
    ("ATR",   True,  None,  True),
    ("sansTP",False, None,  False),
]

# ---------------------------------------------------------------------------
# Cache des DataFrames par année (évite de recalculer les indicateurs)
# ---------------------------------------------------------------------------
_df_year_cache: dict = {}

def get_dfs_for_year(symbols, tf, year):
    key = (tf, year)
    if key not in _df_year_cache:
        dfs = {}
        for sym in symbols:
            df = ms.get_df_for_year(sym, tf, year)
            if df is not None and len(df) >= 10:
                dfs[sym] = df
        _df_year_cache[key] = dfs
    return _df_year_cache[key]

_df_4y_cache: dict = {}

def get_dfs_4y(symbols, tf):
    key = tf
    if key not in _df_4y_cache:
        n = ms.CANDLES_PER_YEAR[tf] * 4
        dfs = {}
        for sym in symbols:
            df = ms.get_df(sym, tf)
            if df is not None and len(df) >= 10:
                dfs[sym] = df.tail(n).reset_index(drop=True)
        _df_4y_cache[key] = dfs
    return _df_4y_cache[key]

# ---------------------------------------------------------------------------
# Simulation helper
# ---------------------------------------------------------------------------
def run_one(dfs, use_sl, strat_name, use_triple_st, use_sma_macd,
            tp_label, use_tp, tp_pct, atr_tp, fg_data, tf):
    if not dfs:
        return None
    try:
        r = ms.sim_multi_on_dfs(
            dfs,
            use_sl=use_sl,
            fg=fg_data,
            use_triple_st=use_triple_st,
            use_sma_macd=use_sma_macd,
            atr_tp=atr_tp,
            use_tp=use_tp,
            tp_pct=tp_pct,
            tf=tf,
        )
        return r.get("return_%")
    except Exception:
        return None

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{Fore.CYAN}{'='*100}")
    print("  TP RANKING — Comparaison des modes de Take-Profit")
    print(f"  {len(TP_MODES)} TP × {len(ALL_TF)} TF × {len(STRATEGIES)} stratégies × 2 SL × {len(YEARS)} années")
    print(f"{'='*100}{Style.RESET_ALL}")

    # ---- Chargement des données ----
    print(f"\n{Fore.CYAN}Chargement des données...{Style.RESET_ALL}")
    data_cache.prefetch_all(SYMBOLS, ALL_TF, verbose=False)

    print(f"{Fore.CYAN}Calcul des indicateurs...{Style.RESET_ALL}")
    total_pairs = len(SYMBOLS) * len(ALL_TF)
    done = 0
    for sym in SYMBOLS:
        for tf in ALL_TF:
            ms.get_df(sym, tf)
            done += 1
            print(f"\r  {done}/{total_pairs}", end="", flush=True)
    print()

    # Précalcul des DFs par année
    print(f"{Fore.CYAN}Précalcul DFs par année...{Style.RESET_ALL}")
    total_year_pairs = len(ALL_TF) * len(YEARS)
    done = 0
    for tf in ALL_TF:
        for year in YEARS:
            get_dfs_for_year(SYMBOLS, tf, year)
            done += 1
            print(f"\r  {done}/{total_year_pairs}", end="", flush=True)
    print()

    print(f"{Fore.CYAN}Fear & Greed...{Style.RESET_ALL}")
    fg_data = fear_greed.load(verbose=False)

    # ---- Simulations années calendaires ----
    total = len(TP_MODES) * len(ALL_TF) * len(STRATEGIES) * 2 * len(YEARS)
    done  = 0
    print(f"\n{Fore.CYAN}Simulations annuelles ({total} backtests)...{Style.RESET_ALL}")

    rows_year = []
    for tp_label, use_tp, tp_pct, atr_tp in TP_MODES:
        for tf in ALL_TF:
            for strat_name, use_triple_st, use_sma_macd in STRATEGIES:
                for use_sl in (True, False):
                    for year in YEARS:
                        dfs = get_dfs_for_year(SYMBOLS, tf, year)
                        ret = run_one(dfs, use_sl, strat_name, use_triple_st,
                                      use_sma_macd, tp_label, use_tp, tp_pct,
                                      atr_tp, fg_data, tf)
                        done += 1
                        if ret is not None:
                            rows_year.append({
                                "tp":       tp_label,
                                "tf":       tf,
                                "stratégie": strat_name,
                                "sl":       "avecSL" if use_sl else "sansSL",
                                "année":    year,
                                "return_%": round(ret, 2),
                            })
                        print(f"\r  {done}/{total}", end="", flush=True)
    print()

    # ---- Simulations 4 ans ----
    total_4y = len(TP_MODES) * len(ALL_TF) * len(STRATEGIES) * 2
    done = 0
    print(f"{Fore.CYAN}Simulations 4 ans ({total_4y} backtests)...{Style.RESET_ALL}")

    rows_4y = []
    for tp_label, use_tp, tp_pct, atr_tp in TP_MODES:
        for tf in ALL_TF:
            for strat_name, use_triple_st, use_sma_macd in STRATEGIES:
                for use_sl in (True, False):
                    dfs = get_dfs_4y(SYMBOLS, tf)
                    ret = run_one(dfs, use_sl, strat_name, use_triple_st,
                                  use_sma_macd, tp_label, use_tp, tp_pct,
                                  atr_tp, fg_data, tf)
                    done += 1
                    if ret is not None:
                        rows_4y.append({
                            "tp":        tp_label,
                            "tf":        tf,
                            "stratégie": strat_name,
                            "sl":        "avecSL" if use_sl else "sansSL",
                            "return_%":  round(ret, 2),
                        })
                    print(f"\r  {done}/{total_4y}", end="", flush=True)
    print()

    # ---- DataFrames ----
    df_year = pd.DataFrame(rows_year)
    df_4y   = pd.DataFrame(rows_4y)

    # ---- Score de régularité par config ----
    print(f"\n{Fore.CYAN}Calcul des scores de régularité...{Style.RESET_ALL}")
    config_cols = ["tp", "tf", "stratégie", "sl"]

    if not df_year.empty:
        grp = df_year.groupby(config_cols)["return_%"]
        reg = pd.DataFrame({
            "n_années":  grp.count(),
            "moy_%":     grp.mean().round(1),
            "min_%":     grp.min().round(1),
            "max_%":     grp.max().round(1),
            "std_%":     grp.std().round(1),
            "n_positif": grp.apply(lambda x: int((x > 0).sum())),
        }).reset_index()
        reg["pct_positif"] = (reg["n_positif"] / reg["n_années"] * 100).round(0).astype(int)
        reg["score"]       = (reg["moy_%"] * (reg["pct_positif"] / 100) ** 2).round(1)
        reg = reg[reg["n_années"] >= 4].sort_values("score", ascending=False).reset_index(drop=True)
        reg.index += 1
    else:
        reg = pd.DataFrame()

    # ---- Résumé par TP (médiane sur toutes configs) ----
    def color(v, fmt="+.1f"):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return "—"
        c = Fore.GREEN if v > 0 else (Fore.RED if v < 0 else "")
        return f"{c}{v:{fmt}}%{Style.RESET_ALL}"

    print(f"\n{Fore.YELLOW}{'='*90}")
    print("  MÉDIANE return% par TP (toutes TF / strats / SL confondues)")
    print(f"{'='*90}{Style.RESET_ALL}")

    for src_label, df_src in [("4 ans", df_4y), ("Années", df_year)]:
        if df_src.empty:
            continue
        med = df_src.groupby("tp")["return_%"].median().reindex([m[0] for m in TP_MODES])
        print(f"\n  {src_label}:")
        for tp_lbl, val in med.items():
            bar = "█" * max(0, int(val / 3)) if not math.isnan(val) and val > 0 else \
                  "░" * max(0, int(-val / 3)) if not math.isnan(val) else ""
            c = Fore.GREEN if not math.isnan(val) and val > 0 else Fore.RED
            print(f"    {tp_lbl:<8} {c}{val:+6.1f}%{Style.RESET_ALL}  {bar}")

    # ---- TOP 30 régularité ----
    if not reg.empty:
        print(f"\n{Fore.CYAN}{'='*110}")
        print("  TOP 30 — Score de régularité  (score = moy% × (pct_positif)²)")
        print(f"{'='*110}{Style.RESET_ALL}")

        reg_rows = []
        for i in range(min(30, len(reg))):
            row = reg.iloc[i]
            reg_rows.append([
                i + 1,
                row["tp"], row["tf"], row["stratégie"], row["sl"],
                row["n_années"],
                f"{Fore.GREEN if row['moy_%'] >= 0 else Fore.RED}{row['moy_%']:+.1f}%{Style.RESET_ALL}",
                f"{Fore.RED if row['min_%'] < 0 else Fore.GREEN}{row['min_%']:+.1f}%{Style.RESET_ALL}",
                f"{row['max_%']:+.1f}%",
                f"{row['std_%']:.1f}",
                f"{row['pct_positif']}%",
                f"{Fore.CYAN}{row['score']:+.1f}{Style.RESET_ALL}",
            ])
        print(tabulate(reg_rows,
                       headers=["#", "TP", "TF", "Stratégie", "SL",
                                 "N_ans", "Moy%", "Min%", "Max%", "Std", "Pct+", "Score"],
                       tablefmt="rounded_outline"))

    # ---- TOP 30 return 4 ans ----
    if not df_4y.empty:
        df_4y_sorted = df_4y.sort_values("return_%", ascending=False).reset_index(drop=True)
        print(f"\n{Fore.YELLOW}{'='*90}")
        print("  TOP 30 — Return% sur 4 ans")
        print(f"{'='*90}{Style.RESET_ALL}")
        top_rows = []
        for i in range(min(30, len(df_4y_sorted))):
            row = df_4y_sorted.iloc[i]
            top_rows.append([
                i + 1,
                row["tp"], row["tf"], row["stratégie"], row["sl"],
                color(row["return_%"]),
            ])
        print(tabulate(top_rows,
                       headers=["#", "TP", "TF", "Stratégie", "SL", "Return 4ans%"],
                       tablefmt="rounded_outline"))

    # ---- Résumé par TP × TF (médiane régularité) ----
    if not reg.empty:
        print(f"\n{Fore.CYAN}{'='*90}")
        print("  SCORE RÉGULARITÉ MÉDIAN par TP × TF")
        print(f"{'='*90}{Style.RESET_ALL}")
        pivot = reg.pivot_table(values="score", index="tp", columns="tf", aggfunc="median")
        pivot = pivot.reindex(index=[m[0] for m in TP_MODES], columns=ALL_TF)
        pivot_rows = []
        for tp_lbl in pivot.index:
            row_data = [tp_lbl]
            for tf in ALL_TF:
                v = pivot.loc[tp_lbl, tf] if tf in pivot.columns else float("nan")
                if math.isnan(v):
                    row_data.append("—")
                else:
                    c = Fore.GREEN if v > 0 else Fore.RED
                    row_data.append(f"{c}{v:+.1f}{Style.RESET_ALL}")
            pivot_rows.append(row_data)
        print(tabulate(pivot_rows, headers=["TP"] + ALL_TF, tablefmt="rounded_outline"))

    # ---- Export CSV ----
    csv_path = "tp_ranking_results.csv"
    if not reg.empty:
        reg.to_csv(csv_path, index=True, float_format="%.2f")
        print(f"\n{Fore.GREEN}Export régularité → {csv_path} ({len(reg)} lignes){Style.RESET_ALL}")

    # ---- Export Excel ----
    xlsx_path = "tp_ranking_results.xlsx"
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GREEN_HEADER = PatternFill("solid", fgColor="C6EFCE")
        GREEN_CELL   = PatternFill("solid", fgColor="EBF5EB")
        RED_HEADER   = PatternFill("solid", fgColor="FFC7CE")
        RED_CELL     = PatternFill("solid", fgColor="FFE7EA")
        GREY_HEADER  = PatternFill("solid", fgColor="D9D9D9")
        BOLD         = Font(bold=True)
        CENTER       = Alignment(horizontal="center")
        thin         = Side(style="thin", color="BBBBBB")
        BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_sheet(ws, df_src, green_cols, red_cols=None):
            red_cols = red_cols or set()
            for col_idx, col_name in enumerate(df_src.columns, start=1):
                max_len = max(len(str(col_name)),
                              df_src[col_name].astype(str).str.len().max())
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 22)
            for col_idx, col_name in enumerate(df_src.columns, start=1):
                if col_name in green_cols:
                    h_fill, c_fill = GREEN_HEADER, GREEN_CELL
                elif col_name in red_cols:
                    h_fill, c_fill = RED_HEADER, RED_CELL
                else:
                    h_fill, c_fill = GREY_HEADER, None
                hc = ws.cell(row=1, column=col_idx)
                hc.fill = h_fill; hc.font = BOLD; hc.alignment = CENTER; hc.border = BORDER
                for row_idx in range(2, len(df_src) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if c_fill:
                        cell.fill = c_fill
                    cell.alignment = CENTER; cell.border = BORDER
            ws.freeze_panes = "A2"

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            if not reg.empty:
                reg_excel = reg.reset_index().rename(columns={"index": "rang"})
                reg_excel.to_excel(writer, index=False, sheet_name="Régularité")
                style_sheet(writer.sheets["Régularité"], reg_excel,
                            green_cols={"moy_%", "score", "pct_positif", "n_positif"},
                            red_cols={"min_%"})

            if not df_4y.empty:
                df4_sorted = df_4y.sort_values("return_%", ascending=False).reset_index(drop=True)
                df4_sorted.index += 1
                df4_excel = df4_sorted.reset_index().rename(columns={"index": "rang"})
                df4_excel.to_excel(writer, index=False, sheet_name="4ans")
                style_sheet(writer.sheets["4ans"], df4_excel,
                            green_cols={"return_%"})

            if not df_year.empty:
                df_year_sorted = df_year.sort_values(["tp","tf","stratégie","sl","année"])
                df_year_sorted.to_excel(writer, index=False, sheet_name="Annuel")
                style_sheet(writer.sheets["Annuel"], df_year_sorted,
                            green_cols={"return_%"})

        print(f"{Fore.GREEN}Export → {xlsx_path} (3 feuilles : Régularité / 4ans / Annuel){Style.RESET_ALL}")
    except ImportError:
        print(f"{Fore.YELLOW}openpyxl non installé — Excel ignoré{Style.RESET_ALL}")
