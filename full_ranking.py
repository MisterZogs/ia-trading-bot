"""
Tableau de classement complet — classement définitif toutes simulations.

Dimensions :
  Portfolios  : top20, top10, top5, btceth, btc, eth
  Modes       : single (1 pos/paire), multi (N pos/paire)
  SL          : avecSL, sansSL
  Timeframes  : 30m, 1h, 2h, 4h, 6h, 12h, 1d
  Périodes    : 1an, 2ans, 3ans, 4ans
  Stratégies  : 4 variantes (baseline, +TripleST, épurée, épurée+ST)

= 6 × 2 × 2 × 7 × 4 × 4 = 2 688 backtests (~8 min)

+ Période ancienne (années -5 à -8) pour TFs 2h/4h/6h/12h/1d
"""

import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

import config
import data_cache
import fear_greed
import multi_sim as ms
from multi_sim import sim_concentration

init(autoreset=True)

# ---------------------------------------------------------------------------
# Dimensions
# ---------------------------------------------------------------------------
ALL_TF     = ["30m", "1h", "2h", "4h", "6h", "12h", "1d"]
OLD_TF     = data_cache.TF_WITH_8Y  # ["2h", "4h", "6h", "12h", "1d"]
PERIODS    = {1: "1an", 2: "2ans", 3: "3ans", 4: "4ans"}
PORTFOLIOS = {
    "top20":  config.SYMBOLS,
    "top10":  ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT",
               "ADA/USDT","AVAX/USDT","DOT/USDT","LINK/USDT","MATIC/USDT"],
    "top5":   ["BTC/USDT","ETH/USDT","BNB/USDT","SOL/USDT","XRP/USDT"],
    "btceth": ["BTC/USDT","ETH/USDT"],
    "btc":    ["BTC/USDT"],
    "eth":    ["ETH/USDT"],
}
PORT_LABELS = {
    "top20":"Top 20","top10":"Top 10","top5":"Top 5",
    "btceth":"BTC+ETH","btc":"BTC","eth":"ETH",
}

# ---------------------------------------------------------------------------
# Concentration du capital : (label, pos_pct, max_trades)
# ---------------------------------------------------------------------------
CAPITAL_SCHEMES = [
    ("5%/illim",  0.05, 9999),
    ("10%/max10", 0.10,   10),
    ("20%/max5",  0.20,    5),
    ("50%/max2",  0.50,    2),
    ("100%/max1", 1.00,    1),
]

# ---------------------------------------------------------------------------
# Stratégies : (nom, use_triple_st, use_sma_macd, min_score)
# ---------------------------------------------------------------------------
STRATEGIES = [
    ("baseline",   False, True,  3),  # 8cond sans TripleST  (référence)
    ("+TripleST",  True,  True,  3),  # 8cond avec TripleST
    ("épurée",     False, False, 3),  # 6cond sans b3/b4 ni TripleST  ← meilleure Top20
    ("épurée+ST",  True,  False, 3),  # 6cond sans b3/b4, avec TripleST
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def buy_and_hold(symbols, tf, years, old_period=False):
    n   = ms.CANDLES_PER_YEAR[tf] * years
    n4y = ms.CANDLES_PER_YEAR[tf] * 4
    returns = []
    for sym in symbols:
        if old_period:
            df = ms.get_df_8y(sym, tf)
        else:
            df = ms.get_df(sym, tf)
        if df is None or len(df) < 10:
            continue
        if old_period:
            if len(df) <= n4y:
                continue
            dp = df.iloc[:-n4y].tail(n)
            if len(dp) < int(n * 0.9):
                continue  # moins de 90% de la période → B&H non représentatif
        else:
            dp = df.tail(n)
        s, e = dp["close"].iloc[0], dp["close"].iloc[-1]
        if s > 0:
            returns.append((e - s) / s * 100)
    return round(sum(returns) / len(returns), 1) if returns else None


CALENDAR_YEARS = list(range(2016, 2027))  # 2016 → 2026

def run_sim(syms, tf, years, mode, use_sl, use_triple_st, use_sma_macd, min_score, fg,
            old_period=False, pos_pct=0.05, max_trades=9999):
    original = config.MIN_SCORE_TO_TRADE
    config.MIN_SCORE_TO_TRADE = min_score
    try:
        # Concentration non-standard : utilise sim_concentration (multi, sansSL par défaut)
        if pos_pct != 0.05 or max_trades != 9999:
            n   = ms.CANDLES_PER_YEAR[tf] * years
            n4y = ms.CANDLES_PER_YEAR[tf] * 4
            dfs = {}
            for sym in syms:
                df = ms.get_df_8y(sym, tf) if old_period else ms.get_df(sym, tf)
                if df is not None and len(df) >= 10:
                    if old_period:
                        if len(df) <= n4y:
                            continue
                        sliced = df.iloc[:-n4y].tail(n).reset_index(drop=True)
                        if len(sliced) < int(n * 0.9):
                            continue
                    else:
                        sliced = df.tail(n).reset_index(drop=True)
                    dfs[sym] = sliced
            r = sim_concentration(dfs, pos_pct=pos_pct, max_trades=max_trades,
                                  fg=fg, use_triple_st=use_triple_st,
                                  use_sma_macd=use_sma_macd,
                                  use_sl=use_sl, tf=tf,
                                  single=(mode == "single"))
        else:
            fn = ms.sim_multi if mode == "multi" else ms.sim_single
            r  = fn(syms, tf, years, use_sl=use_sl, fg=fg,
                    use_triple_st=use_triple_st, use_sma_macd=use_sma_macd,
                    old_period=old_period)
    finally:
        config.MIN_SCORE_TO_TRADE = original
    return r


def run_sim_year(syms, tf, year, mode, use_sl, use_triple_st, use_sma_macd, min_score, fg,
                 pos_pct=0.05, max_trades=9999):
    original = config.MIN_SCORE_TO_TRADE
    config.MIN_SCORE_TO_TRADE = min_score
    try:
        if pos_pct != 0.05 or max_trades != 9999:
            dfs = {}
            for sym in syms:
                df = ms.get_df_for_year(sym, tf, year)
                if df is not None:
                    dfs[sym] = df
            r = sim_concentration(dfs, pos_pct=pos_pct, max_trades=max_trades,
                                  fg=fg, use_triple_st=use_triple_st,
                                  use_sma_macd=use_sma_macd,
                                  use_sl=use_sl, tf=tf,
                                  single=(mode == "single"))
        else:
            r = ms.sim_year(syms, tf, year, mode=mode, use_sl=use_sl, fg=fg,
                            use_triple_st=use_triple_st, use_sma_macd=use_sma_macd)
    finally:
        config.MIN_SCORE_TO_TRADE = original
    return r


def buy_and_hold_year(symbols, tf, year):
    returns = []
    for sym in symbols:
        df = ms.get_df_for_year(sym, tf, year)
        if df is None or len(df) < 10:
            continue
        s, e = df["close"].iloc[0], df["close"].iloc[-1]
        if s > 0:
            returns.append((e - s) / s * 100)
    return round(sum(returns) / len(returns), 1) if returns else None


def color_val(val, fmt="+.1f"):
    if val is None:
        return "—"
    c = Fore.GREEN if val > 0 else (Fore.RED if val < 0 else "")
    return f"{c}{val:{fmt}}%{Style.RESET_ALL}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{Fore.CYAN}{'='*100}")
    print("  CLASSEMENT DÉFINITIF — 4 stratégies × toutes dimensions + période ancienne")
    print(f"{'='*100}{Style.RESET_ALL}")

    all_symbols = list(dict.fromkeys(s for syms in PORTFOLIOS.values() for s in syms))

    # ---- Données récentes (4 ans) -------------------------------------------
    print(f"\n{Fore.CYAN}Chargement données récentes (4 ans)...{Style.RESET_ALL}")
    data_cache.prefetch_all(all_symbols, ALL_TF, verbose=False)

    print(f"{Fore.CYAN}Calcul indicateurs récents...{Style.RESET_ALL}")
    total_pairs = len(all_symbols) * len(ALL_TF)
    done = 0
    for sym in all_symbols:
        for tf in ALL_TF:
            ms.get_df(sym, tf)
            done += 1
            print(f"\r  {done}/{total_pairs}", end="", flush=True)
    print()

    # ---- Données anciennes (8 ans, TFs supportés) ---------------------------
    print(f"{Fore.CYAN}Téléchargement données anciennes (8 ans, {OLD_TF})...{Style.RESET_ALL}")
    data_cache.prefetch_all_8y(all_symbols, OLD_TF, verbose=True)

    print(f"{Fore.CYAN}Calcul indicateurs anciens...{Style.RESET_ALL}")
    total_old = len(all_symbols) * len(OLD_TF)
    done = 0
    for sym in all_symbols:
        for tf in OLD_TF:
            ms.get_df_8y(sym, tf)
            done += 1
            print(f"\r  {done}/{total_old}", end="", flush=True)
    print()

    # ---- Données 10 ans (années calendaires 2016-2017) ----------------------
    print(f"{Fore.CYAN}Téléchargement données 10 ans (années calendaires)...{Style.RESET_ALL}")
    data_cache.prefetch_all_10y(all_symbols, data_cache.TF_WITH_10Y, verbose=True)

    print(f"{Fore.CYAN}Calcul indicateurs 10 ans...{Style.RESET_ALL}")
    total_10y = len(all_symbols) * len(data_cache.TF_WITH_10Y)
    done = 0
    for sym in all_symbols:
        for tf in data_cache.TF_WITH_10Y:
            ms.get_df_10y(sym, tf)
            done += 1
            print(f"\r  {done}/{total_10y}", end="", flush=True)
    print()

    # ---- Fear & Greed -------------------------------------------------------
    print(f"{Fore.CYAN}Fear & Greed...{Style.RESET_ALL}")
    fg_data = fear_greed.load(verbose=False)

    # ---- Buy-and-hold (récent + ancien + années calendaires) ----------------
    print(f"{Fore.CYAN}Buy-and-hold...{Style.RESET_ALL}")
    bah_cache     = {}
    bah_old_cache = {}
    bah_year_cache = {}
    for port, syms in PORTFOLIOS.items():
        for tf in ALL_TF:
            for years in PERIODS:
                bah_cache[(port, tf, years)] = buy_and_hold(syms, tf, years, old_period=False)
        for tf in OLD_TF:
            for years in PERIODS:
                bah_old_cache[(port, tf, years)] = buy_and_hold(syms, tf, years, old_period=True)
        for year in CALENDAR_YEARS:
            for tf in ALL_TF:
                bah_year_cache[(port, tf, year)] = buy_and_hold_year(syms, tf, year)

    # ---- Simulations récentes -----------------------------------------------
    all_rows = []
    total = len(PORTFOLIOS) * 2 * 2 * len(ALL_TF) * len(PERIODS) * len(STRATEGIES) * len(CAPITAL_SCHEMES)
    done  = 0
    print(f"{Fore.CYAN}Simulations récentes ({total} backtests)...{Style.RESET_ALL}")

    for port, syms in PORTFOLIOS.items():
        for mode in ("single", "multi"):
            for use_sl in (True, False):
                for tf in ALL_TF:
                    for years in PERIODS:
                        for strat_name, use_triple_st, use_sma_macd, min_score in STRATEGIES:
                            for cap_label, pos_pct, max_trades in CAPITAL_SCHEMES:
                                # mode single n'a pas de sens avec max_trades > 1 par paire,
                                # mais on conserve pour comparabilité (la limite globale s'applique quand même)
                                try:
                                    r = run_sim(syms, tf, years, mode, use_sl,
                                                use_triple_st, use_sma_macd, min_score, fg_data,
                                                pos_pct=pos_pct, max_trades=max_trades)
                                except Exception:
                                    r = {}
                                bah = bah_cache.get((port, tf, years), 0.0)
                                ret = r.get("return_%")
                                nb_trades = r.get("trades")
                                all_rows.append({
                                    "portfolio":       PORT_LABELS[port],
                                    "mode":            "multi" if mode == "multi" else "1pos",
                                    "sl":              "avecSL" if use_sl else "sansSL",
                                    "timeframe":       tf,
                                    "période":         PERIODS[years],
                                    "stratégie":       strat_name,
                                    "capital":         cap_label,
                                    "return_%":        ret,
                                    "drawdown_%":      r.get("drawdown_%"),
                                    "win_%":           r.get("win_%"),
                                    "trades":          nb_trades,
                                    "pf":              r.get("profit_factor"),
                                    "dur.moy_j":       r.get("avg_duration_j"),
                                    "non_ferm.":       r.get("unclosed"),
                                    "bah_%":           bah,
                                    "alpha_%":         round(ret - bah, 1) if ret is not None and bah is not None else None,
                                    "volume_annuel_%": round(nb_trades * pos_pct / years, 1) if nb_trades else None,
                                    # Colonnes ancien période — remplies après
                                    "ret_old_%":       None,
                                    "dd_old_%":        None,
                                    "bah_old_%":       None,
                                    "alpha_old_%":     None,
                                    "delta_%":         None,
                                })
                                done += 1
                            print(f"\r  {done}/{total}", end="", flush=True)
    print()

    # ---- Simulations anciennes (OLD_TF seulement) ---------------------------
    total_old_sims = len(PORTFOLIOS) * 2 * 2 * len(OLD_TF) * len(PERIODS) * len(STRATEGIES) * len(CAPITAL_SCHEMES)
    done = 0
    print(f"{Fore.CYAN}Simulations anciennes (-5à-8 ans, {total_old_sims} backtests)...{Style.RESET_ALL}")

    # Construire un dict de lookup pour retrouver les rows à compléter
    row_index = {}
    for i, row in enumerate(all_rows):
        key = (row["portfolio"], row["mode"], row["sl"],
               row["timeframe"], row["période"], row["stratégie"], row["capital"])
        row_index[key] = i

    for port, syms in PORTFOLIOS.items():
        for mode in ("single", "multi"):
            for use_sl in (True, False):
                for tf in OLD_TF:
                    for years in PERIODS:
                        for strat_name, use_triple_st, use_sma_macd, min_score in STRATEGIES:
                            for cap_label, pos_pct, max_trades in CAPITAL_SCHEMES:
                                try:
                                    r_old = run_sim(syms, tf, years, mode, use_sl,
                                                    use_triple_st, use_sma_macd, min_score,
                                                    fg_data, old_period=True,
                                                    pos_pct=pos_pct, max_trades=max_trades)
                                except Exception:
                                    r_old = {}
                                bah_old = bah_old_cache.get((port, tf, years))
                                ret_old = r_old.get("return_%")

                                key = (PORT_LABELS[port],
                                       "multi" if mode == "multi" else "1pos",
                                       "avecSL" if use_sl else "sansSL",
                                       tf, PERIODS[years], strat_name, cap_label)
                                if key in row_index:
                                    row = all_rows[row_index[key]]
                                    row["ret_old_%"]   = ret_old
                                    row["dd_old_%"]    = r_old.get("drawdown_%")
                                    row["bah_old_%"]   = bah_old
                                    row["alpha_old_%"] = (
                                        round(ret_old - bah_old, 1)
                                        if ret_old is not None and bah_old is not None else None
                                    )
                                    ret_rec = row["return_%"]
                                    row["delta_%"] = (
                                        round(ret_rec - ret_old, 1)
                                        if ret_rec is not None and ret_old is not None else None
                                    )
                                done += 1
                            print(f"\r  {done}/{total_old_sims}", end="", flush=True)
    print()

    # ---- Simulations années calendaires (2016-2026) -------------------------
    total_year = len(CALENDAR_YEARS) * len(PORTFOLIOS) * 2 * 2 * len(ALL_TF) * len(STRATEGIES) * len(CAPITAL_SCHEMES)
    done = 0
    print(f"{Fore.CYAN}Simulations années calendaires ({total_year} backtests max)...{Style.RESET_ALL}")

    for year in CALENDAR_YEARS:
        for port, syms in PORTFOLIOS.items():
            for mode in ("single", "multi"):
                for use_sl in (True, False):
                    for tf in ALL_TF:
                        for strat_name, use_triple_st, use_sma_macd, min_score in STRATEGIES:
                            for cap_label, pos_pct, max_trades in CAPITAL_SCHEMES:
                                try:
                                    r = run_sim_year(syms, tf, year, mode, use_sl,
                                                     use_triple_st, use_sma_macd, min_score, fg_data,
                                                     pos_pct=pos_pct, max_trades=max_trades)
                                except Exception:
                                    r = {}
                                ret = r.get("return_%")
                                if ret is None:
                                    done += 1
                                    continue
                                bah = bah_year_cache.get((port, tf, year))
                                nb_trades = r.get("trades")
                                all_rows.append({
                                    "portfolio":       PORT_LABELS[port],
                                    "mode":            "multi" if mode == "multi" else "1pos",
                                    "sl":              "avecSL" if use_sl else "sansSL",
                                    "timeframe":       tf,
                                    "période":         str(year),
                                    "stratégie":       strat_name,
                                    "capital":         cap_label,
                                    "return_%":        ret,
                                    "drawdown_%":      r.get("drawdown_%"),
                                    "win_%":           r.get("win_%"),
                                    "trades":          nb_trades,
                                    "pf":              r.get("profit_factor"),
                                    "dur.moy_j":       r.get("avg_duration_j"),
                                    "non_ferm.":       r.get("unclosed"),
                                    "bah_%":           bah,
                                    "alpha_%":         round(ret - bah, 1) if ret is not None and bah is not None else None,
                                    "volume_annuel_%": round(nb_trades * pos_pct, 1) if nb_trades else None,
                                    "ret_old_%":       None,
                                    "dd_old_%":        None,
                                    "bah_old_%":       None,
                                    "alpha_old_%":     None,
                                    "delta_%":         None,
                                })
                                done += 1
                            print(f"\r  {done}/{total_year}", end="", flush=True)
    print()

    # ---- DataFrame trié -----------------------------------------------------
    df = pd.DataFrame(all_rows)
    df = df[df["return_%"].notna()].copy()
    df = df.sort_values("return_%", ascending=False).reset_index(drop=True)
    df.index += 1

    total_valid = len(df)
    print(f"\n{Fore.GREEN}  {total_valid} simulations valides{Style.RESET_ALL}")

    def fmt_row(rank, row):
        ret   = row["return_%"]
        bah   = row["bah_%"]
        alpha = row["alpha_%"]
        dd    = row["drawdown_%"]
        pf    = row["pf"]
        ret_old = row.get("ret_old_%")
        delta   = row.get("delta_%")
        return [
            rank,
            row["portfolio"], row["mode"], row["sl"],
            row["timeframe"], row["période"], row["stratégie"], row.get("capital", "5%/illim"),
            color_val(ret),
            color_val(dd, "+.1f") if dd is not None else "—",
            f"{row['win_%']:.0f}%" if row["win_%"] is not None else "—",
            f"{row['trades']:.0f}" if row["trades"] is not None else "—",
            f"{pf:.2f}" if pf is not None else "—",
            f"{bah:+.1f}%" if bah is not None else "—",
            color_val(alpha),
            color_val(ret_old) if ret_old is not None else "—",
            color_val(delta) if delta is not None else "—",
        ]

    headers = ["#", "Portfolio", "Mode", "SL", "TF", "Période", "Stratégie", "Capital",
               "Return%", "MaxDD%", "Win%", "Trades", "PF", "B&H%", "Alpha",
               "Ret ancien%", "Delta R-A"]

    # TOP 50
    print(f"\n{Fore.YELLOW}{'='*150}")
    print("  TOP 50 — meilleures simulations (récentes)")
    print(f"{'='*150}{Style.RESET_ALL}")
    rows_top = [fmt_row(df.index[i], df.iloc[i]) for i in range(min(50, len(df)))]
    print(tabulate(rows_top, headers=headers, tablefmt="rounded_outline"))

    # BOTTOM 20
    print(f"\n{Fore.RED}{'='*150}")
    print("  BOTTOM 20 — pires simulations")
    print(f"{'='*150}{Style.RESET_ALL}")
    rows_bot = [fmt_row(df.index[i], df.iloc[i]) for i in range(max(0, len(df)-20), len(df))]
    print(tabulate(rows_bot, headers=headers, tablefmt="rounded_outline"))

    # ---- Résumé par dimension -----------------------------------------------
    print(f"\n{Fore.CYAN}{'='*80}")
    print("  RÉSUMÉ PAR DIMENSION (médiane du return%)")
    print(f"{'='*80}{Style.RESET_ALL}")

    for dim, col in [("Stratégie","stratégie"),("Mode","mode"),("SL","sl"),
                     ("Timeframe","timeframe"),("Période","période"),("Portfolio","portfolio"),
                     ("Capital","capital")]:
        grp = df.groupby(col)["return_%"].median().sort_values(ascending=False)
        print(f"\n  {dim}:")
        for k, v in grp.items():
            bar   = "█" * max(0, int(v/5)) if v > 0 else "░" * max(0, int(-v/5))
            color = Fore.GREEN if v > 0 else Fore.RED
            print(f"    {str(k):<22} {color}{v:+6.1f}%{Style.RESET_ALL}  {bar}")

    # ---- Comparaison période récente vs ancienne (OLD_TF, 4ans, multi, sansSL) ----
    print(f"\n{Fore.CYAN}{'='*110}")
    print("  ROBUSTESSE — Récent (0-4ans) vs Ancien (-5à-8ans) | 4ans / multi / sansSL")
    print(f"{'='*110}{Style.RESET_ALL}")

    sub = df[
        (df["timeframe"].isin(OLD_TF)) &
        (df["mode"] == "multi") &
        (df["sl"] == "sansSL") &
        (df["période"] == "4ans") &
        df["ret_old_%"].notna()
    ].copy()

    rob_headers = ["Portfolio", "TF", "Stratégie",
                   "Ret récent%", "Ret ancien%", "Delta",
                   "B&H récent", "B&H ancien",
                   "Alpha récent", "Alpha ancien"]
    rob_rows = []
    for port_label in PORT_LABELS.values():
        for tf in OLD_TF:
            s = sub[(sub["portfolio"] == port_label) & (sub["timeframe"] == tf)]
            s = s.sort_values("return_%", ascending=False)
            if s.empty:
                continue
            best = s.iloc[0]
            rob_rows.append([
                port_label, tf, best["stratégie"],
                color_val(best["return_%"]),
                color_val(best["ret_old_%"]),
                color_val(best["delta_%"]),
                f"{best['bah_%']:+.1f}%" if best["bah_%"] is not None else "—",
                f"{best['bah_old_%']:+.1f}%" if best["bah_old_%"] is not None else "—",
                color_val(best["alpha_%"]),
                color_val(best["alpha_old_%"]),
            ])

    print(tabulate(rob_rows, headers=rob_headers, tablefmt="rounded_outline"))

    # ---- Top par portfolio × stratégie (12h seulement) ----------------------
    print(f"\n{Fore.CYAN}{'='*80}")
    print("  MEILLEURE STRAT PAR PORTFOLIO (12h / multi / sansSL / 4ans)")
    print(f"{'='*80}{Style.RESET_ALL}")

    sub12 = df[(df["timeframe"] == "12h") & (df["mode"] == "multi") &
               (df["sl"] == "sansSL") & (df["période"] == "4ans")].copy()
    sub_rows = []
    for port_label in PORT_LABELS.values():
        s = sub12[sub12["portfolio"] == port_label].sort_values("return_%", ascending=False)
        if s.empty:
            continue
        best = s.iloc[0]
        bah  = best["bah_%"]
        ret  = best["return_%"]
        sub_rows.append([
            port_label,
            best["stratégie"],
            color_val(ret),
            f"{best['drawdown_%']:+.1f}%",
            f"{best['win_%']:.0f}%",
            best["trades"],
            f"{best['pf']:.2f}",
            f"{bah:+.1f}%",
            color_val(best["alpha_%"]),
            color_val(best.get("ret_old_%")),
            color_val(best.get("delta_%")),
        ])

    print(tabulate(sub_rows,
                   headers=["Portfolio","Stratégie","Return%","DD%","Win%","Trades","PF",
                            "B&H%","Alpha","Ret ancien%","Delta"],
                   tablefmt="rounded_outline"))

    # ---- Régularité — performance année après année -------------------------
    print(f"\n{Fore.CYAN}{'='*100}")
    print("  RÉGULARITÉ — configs les plus constantes année après année")
    print(f"  Score = moy_% × (pct_positif/100)²  |  filtre : n_années ≥ 5")
    print(f"{'='*100}{Style.RESET_ALL}")

    year_df = df[df["période"].str.match(r"^\d{4}$", na=False)].copy()

    if not year_df.empty:
        config_cols = ["portfolio", "mode", "sl", "timeframe", "stratégie", "capital"]
        grp = year_df.groupby(config_cols)["return_%"]

        reg_df = pd.DataFrame({
            "n_années":   grp.count(),
            "moy_%":      grp.mean().round(1),
            "min_%":      grp.min().round(1),
            "max_%":      grp.max().round(1),
            "std_%":      grp.std().round(1),
            "n_positif":  year_df.groupby(config_cols)["return_%"].apply(lambda x: int((x > 0).sum())),
        }).reset_index()

        reg_df["pct_positif"] = (reg_df["n_positif"] / reg_df["n_années"] * 100).round(0).astype(int)
        reg_df["score_reg"]   = (reg_df["moy_%"] * (reg_df["pct_positif"] / 100) ** 2).round(1)
        reg_df = reg_df[reg_df["n_années"] >= 5].sort_values("score_reg", ascending=False).reset_index(drop=True)
        reg_df.index += 1

        def fmt_reg_row(rank, row):
            moy = row["moy_%"]
            mn  = row["min_%"]
            return [
                rank,
                row["portfolio"], row["mode"], row["sl"], row["timeframe"], row["stratégie"],
                row.get("capital", "5%/illim"),
                row["n_années"],
                f"{Fore.GREEN if moy >= 0 else Fore.RED}{moy:+.1f}%{Style.RESET_ALL}",
                f"{Fore.RED if mn < 0 else Fore.GREEN}{mn:+.1f}%{Style.RESET_ALL}",
                f"{row['max_%']:+.1f}%",
                f"{row['std_%']:.1f}",
                f"{row['pct_positif']}%",
                f"{Fore.CYAN}{row['score_reg']:+.1f}{Style.RESET_ALL}",
            ]

        reg_headers = ["#", "Portfolio", "Mode", "SL", "TF", "Stratégie", "Capital",
                       "N_ans", "Moy%", "Min%", "Max%", "Std", "Pct+", "Score"]
        reg_rows = [fmt_reg_row(reg_df.index[i], reg_df.iloc[i]) for i in range(min(30, len(reg_df)))]
        print(tabulate(reg_rows, headers=reg_headers, tablefmt="rounded_outline"))
    else:
        reg_df = pd.DataFrame()

    # ---- Export CSV ---------------------------------------------------------
    csv_path = "full_ranking_results.csv"
    df.to_csv(csv_path, index=True, float_format="%.2f")
    print(f"\n{Fore.GREEN}Export → {csv_path} ({total_valid} lignes){Style.RESET_ALL}")

    # ---- Export Excel avec colonnes colorées --------------------------------
    xlsx_path = "full_ranking_results.xlsx"
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        GREEN_HEADER = PatternFill("solid", fgColor="C6EFCE")   # vert clair
        GREEN_CELL   = PatternFill("solid", fgColor="EBF5EB")   # vert très clair
        BLUE_HEADER  = PatternFill("solid", fgColor="BDD7EE")   # bleu clair
        BLUE_CELL    = PatternFill("solid", fgColor="EBF3FB")   # bleu très clair
        GREY_HEADER  = PatternFill("solid", fgColor="D9D9D9")   # gris pour les colonnes meta
        BOLD         = Font(bold=True)
        CENTER       = Alignment(horizontal="center")
        thin         = Side(style="thin", color="BBBBBB")
        BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

        GREEN_COLS = {"return_%", "drawdown_%", "win_%", "trades", "pf", "bah_%", "alpha_%", "capital"}
        BLUE_COLS  = {"ret_old_%", "dd_old_%", "bah_old_%", "alpha_old_%"}

        df_excel = df.reset_index()  # amène le rang (index) comme colonne
        df_excel = df_excel.rename(columns={"index": "rang"})

        def _style_sheet(ws, df_src, green_set, blue_set):
            """Applique largeur, couleurs, alignement et gel sur une feuille."""
            for col_idx, col_name in enumerate(df_src.columns, start=1):
                max_len = max(len(str(col_name)),
                              df_src[col_name].astype(str).str.len().max())
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 22)

            for col_idx, col_name in enumerate(df_src.columns, start=1):
                if col_name in green_set:
                    h_fill, c_fill = GREEN_HEADER, GREEN_CELL
                elif col_name in blue_set:
                    h_fill, c_fill = BLUE_HEADER, BLUE_CELL
                else:
                    h_fill, c_fill = GREY_HEADER, None

                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.fill      = h_fill
                header_cell.font      = BOLD
                header_cell.alignment = CENTER
                header_cell.border    = BORDER

                for row_idx in range(2, len(df_src) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if c_fill:
                        cell.fill = c_fill
                    cell.alignment = CENTER
                    cell.border    = BORDER

            ws.freeze_panes = "A2"

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            # Feuille 1 — Ranking
            df_excel.to_excel(writer, index=False, sheet_name="Ranking")
            _style_sheet(writer.sheets["Ranking"], df_excel, GREEN_COLS, BLUE_COLS)

            # Feuille 2 — Régularité
            if not reg_df.empty:
                reg_excel = reg_df.reset_index().rename(columns={"index": "rang"})
                reg_excel.to_excel(writer, index=False, sheet_name="Régularité")
                REG_GREEN = {"moy_%", "score_reg", "pct_positif", "n_positif"}
                REG_RED   = {"min_%"}
                _style_sheet(writer.sheets["Régularité"], reg_excel, REG_GREEN, REG_RED)

        print(f"{Fore.GREEN}Export → {xlsx_path} ({total_valid} lignes, 2 feuilles){Style.RESET_ALL}")
    except ImportError:
        print(f"{Fore.YELLOW}openpyxl non installé — Excel ignoré (pip install openpyxl){Style.RESET_ALL}")
