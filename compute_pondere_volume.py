"""
Calcule le volume_annuel_% réel pour les lignes pondéré / pondéré-strict
en relançant les simulations et en trackant pos_pct à chaque BUY effectif.

Résultat : met à jour la colonne volume_annuel_% dans full_ranking_results.csv/.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import config
import data_cache
import fear_greed
import multi_sim as ms
from multi_sim import sim_multi_weighted
from etape19_capital_pondere import (
    PORTFOLIOS, PORT_LABELS, STRATEGIES, WEIGHT_VARIANTS,
    SCORE_YEARS, MIN_YEARS, compute_weights,
    run_weighted_period, run_weighted_year,
)

ALL_TF         = ["30m", "1h", "2h", "4h", "6h", "12h", "1d"]
PERIODS        = {1: "1an", 2: "2ans", 3: "3ans", 4: "4ans"}
CALENDAR_YEARS = list(range(2016, 2027))


def period_to_years(p):
    if p == '4ans': return 4
    if p == '3ans': return 3
    if p == '2ans': return 2
    if p == '1an':  return 1
    if str(p).isdigit(): return 1
    return 1


if __name__ == "__main__":
    # ---- Chargement données ------------------------------------------------
    all_symbols = list(dict.fromkeys(s for syms in PORTFOLIOS.values() for s in syms))
    print("Chargement données...")
    data_cache.prefetch_all(all_symbols, ALL_TF, verbose=False)
    for sym in all_symbols:
        for tf in ALL_TF:
            ms.get_df(sym, tf)
    fg = fear_greed.load()

    # ---- Pré-calcul des poids par (portfolio, tf, stratégie, variante) -----
    # Clé : (port_key, tf, strat_name, floor_pct) → weights dict
    weights_cache = {}

    total_combos = len(PORTFOLIOS) * len(ALL_TF) * len(STRATEGIES) * len(WEIGHT_VARIANTS)
    done = 0
    print(f"Calcul des poids de régularité ({total_combos} combinaisons)...")

    for port_key, symbols in PORTFOLIOS.items():
        for tf in ALL_TF:
            for strat_name, use_triple_st, use_sma_macd, _ in STRATEGIES:
                for variant_name, floor_pct in WEIGHT_VARIANTS:
                    key = (port_key, tf, strat_name, floor_pct)
                    if key not in weights_cache:
                        weights_cache[key] = compute_weights(
                            symbols, tf, use_triple_st, use_sma_macd, fg, floor_pct
                        )
                    done += 1
                    print(f"\r  {done}/{total_combos} poids calculés", end="", flush=True)
    print()

    # ---- Simulations et extraction de total_pos_pct ------------------------
    # Structure : (port_label, tf, strat_name, variant_name, période) → total_pos_pct, nb_trades
    results = {}

    # Périodes glissantes
    for port_key, symbols in PORTFOLIOS.items():
        port_label = PORT_LABELS[port_key]
        for tf in ALL_TF:
            for strat_name, use_triple_st, use_sma_macd, _ in STRATEGIES:
                for variant_name, floor_pct in WEIGHT_VARIANTS:
                    key_w = (port_key, tf, strat_name, floor_pct)
                    weights = weights_cache[key_w]
                    for nb_years, period_label in PERIODS.items():
                        for use_sl in [False, True]:
                            sl_label = "avecSL" if use_sl else "sansSL"
                            r = run_weighted_period(
                                symbols, tf, nb_years, weights,
                                use_sl, use_triple_st, use_sma_macd, fg
                            )
                            if r:
                                k = (port_label, sl_label, tf, strat_name, variant_name, period_label)
                                results[k] = (r.get("total_pos_pct", 0.0), r.get("trades", 0), nb_years)

    # Années calendaires
    for port_key, symbols in PORTFOLIOS.items():
        port_label = PORT_LABELS[port_key]
        for tf in ALL_TF:
            for strat_name, use_triple_st, use_sma_macd, _ in STRATEGIES:
                for variant_name, floor_pct in WEIGHT_VARIANTS:
                    key_w = (port_key, tf, strat_name, floor_pct)
                    weights = weights_cache[key_w]
                    for year in CALENDAR_YEARS:
                        for use_sl in [False, True]:
                            sl_label = "avecSL" if use_sl else "sansSL"
                            r = run_weighted_year(
                                symbols, tf, year, weights,
                                use_sl, use_triple_st, use_sma_macd, fg
                            )
                            if r:
                                k = (port_label, sl_label, tf, strat_name, variant_name, str(year))
                                results[k] = (r.get("total_pos_pct", 0.0), r.get("trades", 0), 1)

    print(f"\n{len(results)} simulations complètes.")

    # ---- Mise à jour du CSV ------------------------------------------------
    df = pd.read_csv('/Users/gaetan/Documents/IA/ia-trading-bot/full_ranking_results.csv', index_col=0)

    def update_volume(row):
        if row['capital'] not in ('pondéré', 'pondéré-strict'):
            return row['volume_annuel_%']
        variant_name = row['capital']
        période = str(row['période'])
        nb_years = period_to_years(période)
        k = (row['portfolio'], row['sl'], row['timeframe'], row['stratégie'], variant_name, période)
        if k in results:
            total_pct, nb_trades, _ = results[k]
            if nb_years > 0:
                return round(total_pct / nb_years, 1)
        return row['volume_annuel_%']

    df['volume_annuel_%'] = df.apply(update_volume, axis=1)

    # ---- Sauvegarde --------------------------------------------------------
    df.to_csv('/Users/gaetan/Documents/IA/ia-trading-bot/full_ranking_results.csv')

    path_xlsx = '/Users/gaetan/Documents/IA/ia-trading-bot/full_ranking_results.xlsx'
    df.to_excel(path_xlsx, index=True)

    wb = load_workbook(path_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_comment = headers.index('commentaire') + 1 if 'commentaire' in headers else None
    col_vol     = headers.index('volume_annuel_%') + 1

    color_map = {
        '🏆': 'FFD700', '🥇': 'FFA500', '🥈': 'C0C0C0',
        '⭐': 'ADD8E6', '✅': 'D4EDDA', '💎': 'E8D5F5',
        '🚀': 'FFE4B5', '📊': 'DDEEFF',
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if col_comment:
            c = row[col_comment - 1]
            val = str(c.value or '')
            if val:
                for emoji, color in color_map.items():
                    if val.startswith(emoji):
                        c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        c.font = Font(bold=True)
                        break
        v = row[col_vol - 1]
        if v.value is not None:
            v.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            v.alignment = Alignment(horizontal='right')

    if col_comment:
        ws.column_dimensions[get_column_letter(col_comment)].width = 60
    ws.column_dimensions[get_column_letter(col_vol)].width = 18
    wb.save(path_xlsx)

    na_left = df['volume_annuel_%'].isna().sum()
    print(f"Sauvegardé. NA restants : {na_left}")
